from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
import os
from dotenv import load_dotenv
import psycopg2
from psycopg2.extras import RealDictCursor
from werkzeug.security import generate_password_hash, check_password_hash
from apscheduler.schedulers.background import BackgroundScheduler
from datetime import datetime, date, timedelta
from functools import wraps
from io import BytesIO
import smtplib
from email.message import EmailMessage
import requests
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image

from openpyxl import load_workbook

load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "chave_local_dev_troque_em_producao")
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("FLASK_ENV") == "production"

DATABASE_URL = os.environ.get("DATABASE_URL")
ALERTA_DIAS = 15
ESTOQUES = {
    "almoxarifado": "Almoxarifado",
    "farmacia_satelite": "Farmácia Satélite",
    "carrinho_urgencia": "Carrinho de Urgência",
}


def dias_alerta_vencimento(tipo_estoque):
    return 30 if tipo_estoque == "carrinho_urgencia" else 15


def conectar():
    if not DATABASE_URL:
        raise RuntimeError(
            "DATABASE_URL não configurada. Configure no .env local ou nas variáveis do Render."
        )

    return psycopg2.connect(
        DATABASE_URL,
        cursor_factory=RealDictCursor,
        connect_timeout=10
    )


def criar_banco():
    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS admin (
            id SERIAL PRIMARY KEY,
            usuario TEXT NOT NULL UNIQUE,
            senha_hash TEXT NOT NULL
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS categorias (
            id SERIAL PRIMARY KEY,
            nome TEXT NOT NULL UNIQUE,
            descricao TEXT,
            ativo BOOLEAN NOT NULL DEFAULT TRUE,
            criado_em TIMESTAMP NOT NULL DEFAULT NOW()
        )
    """)
    cursor.execute("ALTER TABLE categorias ADD COLUMN IF NOT EXISTS ativo BOOLEAN NOT NULL DEFAULT TRUE")

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS produtos (
            id SERIAL PRIMARY KEY,
            nome TEXT NOT NULL,
            categoria_id INTEGER NOT NULL REFERENCES categorias(id),
            codigo_barras TEXT,
            fabricante TEXT,
            unidade_medida TEXT,
            validade_apos_aberto_dias INTEGER,
            observacoes TEXT,
            ativo BOOLEAN NOT NULL DEFAULT TRUE,
            criado_em TIMESTAMP NOT NULL DEFAULT NOW()
        )
    """)

    cursor.execute("ALTER TABLE produtos ADD COLUMN IF NOT EXISTS codigo_barras TEXT")
    cursor.execute("ALTER TABLE produtos ADD COLUMN IF NOT EXISTS fabricante TEXT")
    cursor.execute("ALTER TABLE produtos ADD COLUMN IF NOT EXISTS unidade_medida TEXT")
    cursor.execute("ALTER TABLE produtos ADD COLUMN IF NOT EXISTS ativo BOOLEAN NOT NULL DEFAULT TRUE")

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS produto_estoques (
            id SERIAL PRIMARY KEY,
            produto_id INTEGER NOT NULL REFERENCES produtos(id) ON DELETE CASCADE,
            tipo_estoque TEXT NOT NULL,
            estoque_padrao INTEGER NOT NULL DEFAULT 0,
            limite_alerta INTEGER NOT NULL DEFAULT 0,
            ativo BOOLEAN NOT NULL DEFAULT TRUE,
            criado_em TIMESTAMP NOT NULL DEFAULT NOW(),
            UNIQUE (produto_id, tipo_estoque)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS lotes (
            id SERIAL PRIMARY KEY,
            produto_estoque_id INTEGER NOT NULL REFERENCES produto_estoques(id) ON DELETE CASCADE,
            numero_lote TEXT,
            data_vencimento DATE NOT NULL,
            data_abertura DATE,
            quantidade_inicial INTEGER NOT NULL DEFAULT 0,
            quantidade_atual INTEGER NOT NULL DEFAULT 0,
            data_entrada TIMESTAMP NOT NULL DEFAULT NOW(),
            observacoes TEXT,
            ativo BOOLEAN NOT NULL DEFAULT TRUE
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS produtos_abertos (
            id SERIAL PRIMARY KEY,
            lote_id INTEGER REFERENCES lotes(id),
            produto_id INTEGER REFERENCES produtos(id),
            produto_nome TEXT NOT NULL,
            tipo_estoque TEXT NOT NULL,
            numero_lote TEXT,
            data_abertura DATE NOT NULL,
            validade_apos_aberto_dias INTEGER NOT NULL,
            vencimento_apos_aberto DATE NOT NULL,
            quantidade INTEGER NOT NULL DEFAULT 1,
            usuario_id INTEGER,
            usuario_nome TEXT,
            observacoes TEXT,
            criado_em TIMESTAMP NOT NULL DEFAULT NOW()
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS movimentacoes (
            id SERIAL PRIMARY KEY,
            produto_id INTEGER,
            produto_nome TEXT NOT NULL,
            tipo_movimentacao TEXT NOT NULL,
            quantidade INTEGER NOT NULL,
            estoque_anterior INTEGER NOT NULL,
            estoque_atual INTEGER NOT NULL,
            observacao TEXT,
            tipo_estoque TEXT DEFAULT 'almoxarifado',
            estoque_origem TEXT,
            estoque_destino TEXT,
            usuario_id INTEGER,
            usuario_nome TEXT,
            data_movimentacao TIMESTAMP NOT NULL DEFAULT NOW()
        )
    """)

    cursor.execute("ALTER TABLE movimentacoes ADD COLUMN IF NOT EXISTS tipo_estoque TEXT DEFAULT 'almoxarifado'")
    cursor.execute("ALTER TABLE movimentacoes ADD COLUMN IF NOT EXISTS estoque_origem TEXT")
    cursor.execute("ALTER TABLE movimentacoes ADD COLUMN IF NOT EXISTS estoque_destino TEXT")
    cursor.execute("ALTER TABLE movimentacoes ADD COLUMN IF NOT EXISTS usuario_id INTEGER")
    cursor.execute("ALTER TABLE movimentacoes ADD COLUMN IF NOT EXISTS usuario_nome TEXT")
    cursor.execute("ALTER TABLE movimentacoes ADD COLUMN IF NOT EXISTS lote_id INTEGER")
    cursor.execute("ALTER TABLE movimentacoes ADD COLUMN IF NOT EXISTS quantidade_anterior INTEGER")
    cursor.execute("ALTER TABLE movimentacoes ADD COLUMN IF NOT EXISTS quantidade_posterior INTEGER")
    cursor.execute("ALTER TABLE movimentacoes ADD COLUMN IF NOT EXISTS origem TEXT")
    cursor.execute("ALTER TABLE movimentacoes ADD COLUMN IF NOT EXISTS destino TEXT")
    cursor.execute("ALTER TABLE movimentacoes ADD COLUMN IF NOT EXISTS motivo TEXT")
    cursor.execute("ALTER TABLE movimentacoes ADD COLUMN IF NOT EXISTS criado_em TIMESTAMP NOT NULL DEFAULT NOW()")


    cursor.execute("""
        CREATE TABLE IF NOT EXISTS licencas (
            id SERIAL PRIMARY KEY,
            empresa TEXT NOT NULL,
            plano TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'ativo',
            data_vencimento DATE NOT NULL,
            dias_carencia INTEGER NOT NULL DEFAULT 5,
            criado_em TIMESTAMP NOT NULL DEFAULT NOW()
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS configuracoes_alerta (
            id SERIAL PRIMARY KEY,
            email_destino TEXT,
            telefone_whatsapp TEXT,
            usar_email BOOLEAN NOT NULL DEFAULT FALSE,
            usar_whatsapp BOOLEAN NOT NULL DEFAULT FALSE,
            alertar_vencimentos BOOLEAN NOT NULL DEFAULT TRUE,
            alertar_estoque BOOLEAN NOT NULL DEFAULT TRUE,
            alertar_ordem_compra BOOLEAN NOT NULL DEFAULT TRUE,
            hora_envio TEXT NOT NULL DEFAULT '08:00',
            criado_em TIMESTAMP NOT NULL DEFAULT NOW()
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS historico_alertas (
            id SERIAL PRIMARY KEY,
            tipo_alerta TEXT NOT NULL,
            canal TEXT NOT NULL,
            destino TEXT,
            conteudo TEXT,
            status TEXT NOT NULL,
            erro TEXT,
            criado_em TIMESTAMP NOT NULL DEFAULT NOW()
        )
    """)

    cursor.execute("SELECT id FROM admin WHERE usuario = %s", ("admin",))
    admin = cursor.fetchone()

    if not admin:
        cursor.execute(
            "INSERT INTO admin (usuario, senha_hash) VALUES (%s, %s)",
            ("admin", generate_password_hash("admin123"))
        )

    cursor.execute("ALTER TABLE licencas ADD COLUMN IF NOT EXISTS chave_licenca TEXT")
    cursor.execute("ALTER TABLE licencas ADD COLUMN IF NOT EXISTS observacoes TEXT")
    cursor.execute("ALTER TABLE licencas ADD COLUMN IF NOT EXISTS atualizado_em TIMESTAMP DEFAULT NOW()")

    cursor.execute("SELECT id FROM licencas LIMIT 1")
    licenca = cursor.fetchone()

    if not licenca:
        data_vencimento = date.today() + timedelta(days=30)
        cursor.execute("""
            INSERT INTO licencas (empresa, plano, status, data_vencimento, dias_carencia, chave_licenca, observacoes)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            "Hospital Oftalmológico",
            "Premium Mensal",
            "ativo",
            data_vencimento,
            5,
            "HO-PREMIUM-MENSAL-001",
            "Licença inicial gerada automaticamente."
        ))

    cursor.execute("ALTER TABLE configuracoes_alerta ADD COLUMN IF NOT EXISTS intervalo_minutos INTEGER NOT NULL DEFAULT 720")
    cursor.execute("ALTER TABLE configuracoes_alerta ADD COLUMN IF NOT EXISTS ultimo_envio_whatsapp TIMESTAMP")

    cursor.execute("SELECT id FROM configuracoes_alerta LIMIT 1")
    config_alerta = cursor.fetchone()

    if not config_alerta:
        cursor.execute("""
            INSERT INTO configuracoes_alerta (
                email_destino,
                telefone_whatsapp,
                intervalo_minutos,
                usar_email,
                usar_whatsapp,
                alertar_vencimentos,
                alertar_estoque,
                alertar_ordem_compra,
                hora_envio
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            "",
            "",
            720,
            False,
            False,
            True,
            True,
            True,
            "08:00"
        ))

    conn.commit()
    conn.close()




def garantir_tabela_usuarios():
    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id SERIAL PRIMARY KEY,
            nome TEXT NOT NULL,
            usuario TEXT UNIQUE NOT NULL,
            senha_hash TEXT NOT NULL,
            perfil TEXT NOT NULL DEFAULT 'estoque',
            ativo BOOLEAN NOT NULL DEFAULT TRUE,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    cursor.execute("SELECT COUNT(*) AS total FROM usuarios")
    total = cursor.fetchone()["total"]

    if total == 0:
        cursor.execute("""
            INSERT INTO usuarios (nome, usuario, senha_hash, perfil, ativo)
            VALUES (%s, %s, %s, %s, TRUE)
        """, (
            "Sala Central",
            "admin",
            generate_password_hash("admin123"),
            "admin"
        ))

        cursor.execute("""
            INSERT INTO usuarios (nome, usuario, senha_hash, perfil, ativo)
            VALUES (%s, %s, %s, %s, TRUE)
        """, (
            "Estoque",
            "estoque",
            generate_password_hash("admin123"),
            "estoque"
        ))

    conn.commit()
    conn.close()


def buscar_usuario_login(usuario):
    garantir_tabela_usuarios()

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT *
        FROM usuarios
        WHERE usuario = %s
          AND ativo = TRUE
        LIMIT 1
    """, (usuario,))
    user = cursor.fetchone()
    conn.close()
    return user


def listar_usuarios():
    garantir_tabela_usuarios()

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, nome, usuario, perfil, ativo, criado_em
        FROM usuarios
        ORDER BY nome
    """)
    usuarios = cursor.fetchall()
    conn.close()
    return usuarios


def perfil_atual():
    return session.get("perfil", "admin")


def admin_obrigatorio(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if session.get("perfil", "admin") != "admin":
            flash("Acesso restrito à Sala Central.", "erro")
            return redirect(url_for("produtos"))
        return func(*args, **kwargs)
    return wrapper


def estoque_obrigatorio(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if session.get("perfil") not in ["admin", "estoque"]:
            flash("Acesso não autorizado.", "erro")
            return redirect(url_for("login"))
        return func(*args, **kwargs)
    return wrapper


def login_obrigatorio(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if not session.get("logado"):
            return redirect(url_for("login"))
        return func(*args, **kwargs)
    return wrapper


def obter_licenca():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM licencas ORDER BY id LIMIT 1")
    licenca = cursor.fetchone()
    conn.close()
    return licenca


def situacao_licenca():
    licenca = obter_licenca()

    if not licenca:
        return {
            "situacao": "bloqueada",
            "licenca": None,
            "dias_restantes": 0,
            "mensagem": "Licença não encontrada."
        }

    hoje = date.today()
    data_vencimento = converter_data(licenca["data_vencimento"])
    dias_carencia = licenca["dias_carencia"] or 5

    if licenca["status"] == "bloqueado":
        return {
            "situacao": "bloqueada",
            "licenca": licenca,
            "dias_restantes": 0,
            "mensagem": "Licença bloqueada pelo suporte."
        }

    dias_restantes = (data_vencimento - hoje).days

    if licenca["status"] == "ativo" and dias_restantes >= 0:
        return {
            "situacao": "ativa",
            "licenca": licenca,
            "dias_restantes": dias_restantes,
            "mensagem": "Licença ativa."
        }

    dias_vencida = (hoje - data_vencimento).days

    if dias_vencida <= dias_carencia:
        return {
            "situacao": "vencida_limitada",
            "licenca": licenca,
            "dias_restantes": dias_carencia - dias_vencida,
            "mensagem": "Licença vencida. Sistema em modo consulta."
        }

    return {
        "situacao": "bloqueada",
        "licenca": licenca,
        "dias_restantes": 0,
        "mensagem": "Licença vencida e fora do período de carência."
    }


def licenca_obrigatoria(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        info = situacao_licenca()

        if info["situacao"] == "bloqueada":
            return redirect(url_for("licenca_bloqueada"))

        return func(*args, **kwargs)
    return wrapper


def alteracao_permitida(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        info = situacao_licenca()

        if info["situacao"] == "bloqueada":
            return redirect(url_for("licenca_bloqueada"))

        if info["situacao"] == "vencida_limitada":
            flash("Licença vencida. O sistema está em modo consulta e não permite alterações.", "erro")
            return redirect(request.referrer or url_for("dashboard"))

        return func(*args, **kwargs)
    return wrapper


def converter_data(valor):
    if not valor:
        return None
    if isinstance(valor, date):
        return valor
    return datetime.strptime(str(valor), "%Y-%m-%d").date()


def formatar_data(valor):
    if not valor:
        return "-"
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y %H:%M")
    if isinstance(valor, date):
        return valor.strftime("%d/%m/%Y")
    return str(valor)


def nome_tipo_estoque(tipo):
    return ESTOQUES.get(tipo, "Almoxarifado")


@app.context_processor
def contexto_global():
    return {
        "estoques_sistema": ESTOQUES,
        "nome_tipo_estoque": nome_tipo_estoque,
    }



def produto_esta_vencido(produto):
    hoje = date.today()

    data_vencimento = converter_data(produto["data_vencimento"])
    if data_vencimento and data_vencimento < hoje:
        return True

    data_abertura = converter_data(produto["data_abertura"])
    validade_dias = produto["validade_apos_aberto_dias"]

    if data_abertura and validade_dias:
        vencimento_apos_aberto = data_abertura + timedelta(days=int(validade_dias))
        if vencimento_apos_aberto < hoje:
            return True

    return False

def calcular_status(produto):
    hoje = date.today()
    alertas = []
    dias_alerta = dias_alerta_vencimento(produto.get("tipo_estoque"))

    data_vencimento = converter_data(produto["data_vencimento"])

    if data_vencimento:
        dias = (data_vencimento - hoje).days

        if dias < 0:
            alertas.append({"tipo": "danger", "prioridade": 2, "texto": "Vencido pela validade original"})
        elif dias <= dias_alerta:
            alertas.append({"tipo": "warning", "prioridade": 4, "texto": f"Vence em {dias} dia(s)"})

    data_abertura = converter_data(produto["data_abertura"])
    validade_dias = produto["validade_apos_aberto_dias"]
    vencimento_apos_aberto = None

    if data_abertura and validade_dias:
        vencimento_apos_aberto = data_abertura + timedelta(days=int(validade_dias))
        dias_aberto = (vencimento_apos_aberto - hoje).days

        if dias_aberto < 0:
            alertas.append({"tipo": "danger", "prioridade": 1, "texto": "Vencido após abertura"})
        elif dias_aberto <= dias_alerta:
            alertas.append({"tipo": "warning", "prioridade": 3, "texto": f"Vence após aberto em {dias_aberto} dia(s)"})

    quantidade = produto["quantidade_atual"]
    limite = produto["limite_alerta"]

    if quantidade == 0:
        alertas.append({"tipo": "danger", "prioridade": 5, "texto": "Estoque zerado"})
    elif quantidade <= limite:
        alertas.append({"tipo": "stock", "prioridade": 6, "texto": "Estoque baixo"})

    if not alertas:
        return {"tipo": "success", "texto": "Produto OK", "vencimento_apos_aberto": vencimento_apos_aberto}

    principal = sorted(alertas, key=lambda x: x["prioridade"])[0]
    return {"tipo": principal["tipo"], "texto": principal["texto"], "vencimento_apos_aberto": vencimento_apos_aberto}



# =========================================================
# BACKUP AUTOMATICO DO SISTEMA
# =========================================================

BACKUP_DIR = os.path.join(os.getcwd(), "backups")

def garantir_pasta_backup():
    os.makedirs(BACKUP_DIR, exist_ok=True)

def serializar_linhas(linhas):
    resultado = []

    for item in linhas:
        novo = {}

        for chave, valor in dict(item).items():
            if isinstance(valor, (datetime, date)):
                novo[chave] = valor.isoformat()
            else:
                novo[chave] = valor

        resultado.append(novo)

    return resultado


def enviar_backup_para_supabase(caminho_arquivo):
    supabase_url = os.environ.get("SUPABASE_URL")
    service_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    bucket = os.environ.get("SUPABASE_BACKUP_BUCKET", "backups")

    if not supabase_url or not service_key:
        return False, "Supabase Storage não configurado."

    nome_arquivo = os.path.basename(caminho_arquivo)
    storage_path = f"automaticos/{nome_arquivo}"

    bucket = bucket.strip().strip("/")
    storage_path = storage_path.strip().lstrip("/")
    
    supabase_url = supabase_url.strip().rstrip("/")
    supabase_url = supabase_url.replace("/rest/v1", "").replace("/storage/v1", "")

    url = f"{supabase_url.rstrip('/')}/storage/v1/object/{bucket}/{storage_path}"

    headers = {
        "Authorization": f"Bearer {service_key}",
        "apikey": service_key,
        "Content-Type": "application/json",
        "x-upsert": "true"
    }

    with open(caminho_arquivo, "rb") as arquivo:
        resposta = requests.post(url, headers=headers, data=arquivo, timeout=60)

    if resposta.status_code not in [200, 201]:
        return False, f"Erro Supabase Storage: {resposta.status_code} - {resposta.text}"

    return True, f"Backup enviado para Supabase Storage: {storage_path}"

def gerar_backup_sistema():
    garantir_pasta_backup()

    conn = conectar()
    cursor = conn.cursor()

    tabelas = [
        "produtos",
        "produto_estoques",
        "lotes",
        "movimentacoes",
        "usuarios",
        "categorias",
        "ordens_compra",
        "licencas",
        "configuracoes_alerta"
    ]

    backup = {
        "gerado_em": datetime.now().isoformat(),
        "sistema": "Controle Oftalmológico Premium",
        "dados": {}
    }

    for tabela in tabelas:
        try:
            cursor.execute(f"SELECT * FROM {tabela}")
            linhas = cursor.fetchall()
            backup["dados"][tabela] = serializar_linhas(linhas)
        except Exception:
            backup["dados"][tabela] = []

    conn.close()


    nome = datetime.now().strftime("Backup_Auto_App_%d-%m-%Y_%Hh%Mm.json")

    nome = datetime.now().strftime(
    "Backup_Completo_%d-%m-%Y_%Hh%Mm.json"
)

    caminho = os.path.join(BACKUP_DIR, nome)

    with open(caminho, "w", encoding="utf-8") as arquivo:
        json.dump(backup, arquivo, ensure_ascii=False, indent=4)

    upload_ok = False
    upload_mensagem = "Upload externo não executado."

    try:
        upload_ok, upload_mensagem = enviar_backup_para_supabase(caminho)
    except Exception as erro:
        upload_ok = False
        upload_mensagem = f"Erro ao enviar backup externo: {erro}"

    limpar_backups_antigos()

    return caminho, upload_ok, upload_mensagem

def limpar_backups_antigos(limite=30):
    garantir_pasta_backup()

    arquivos = sorted(
        [os.path.join(BACKUP_DIR, f) for f in os.listdir(BACKUP_DIR) if f.endswith(".json")],
        key=os.path.getmtime,
        reverse=True
    )

    for antigo in arquivos[limite:]:
        try:
            os.remove(antigo)
        except Exception:
            pass

def iniciar_backup_automatico():
    # Evita criar múltiplos agendadores no Flask debug/reloader
    if os.environ.get("WERKZEUG_RUN_MAIN") == "false":
        return

    scheduler = BackgroundScheduler(daemon=True)

    scheduler.add_job(
        gerar_backup_sistema,
        "cron",
        hour=5,
        minute=0,
        id="backup_diario_sistema",
        replace_existing=True
    )

    scheduler.start()

@app.route("/painel_master_backup_9182")
@login_obrigatorio
@admin_obrigatorio
def painel_backup():
    garantir_pasta_backup()

    backups = []

    for arquivo in sorted(os.listdir(BACKUP_DIR), reverse=True):
        if arquivo.endswith(".json"):
            caminho = os.path.join(BACKUP_DIR, arquivo)

            backups.append({
                "nome": arquivo,
                "tamanho": round(os.path.getsize(caminho) / 1024, 2),
                "modificado": datetime.fromtimestamp(os.path.getmtime(caminho))
            })

    return render_template(
        "backup_sistema.html",
        backups=backups,
        supabase_configurado=bool(os.environ.get("SUPABASE_URL") and os.environ.get("SUPABASE_SERVICE_ROLE_KEY")),
        supabase_bucket=os.environ.get("SUPABASE_BACKUP_BUCKET", "backups")
    )


@app.route("/backup/testar_supabase", methods=["POST"])
@login_obrigatorio
@admin_obrigatorio
def testar_backup_supabase():
    caminho, upload_ok, upload_mensagem = gerar_backup_sistema()

    flash(upload_mensagem, "sucesso" if upload_ok else "erro")
    return redirect(url_for("painel_backup"))

@app.route("/backup/manual", methods=["POST"])
@login_obrigatorio
@admin_obrigatorio
def gerar_backup_manual():
    gerar_backup_sistema()
    flash("Backup manual gerado com sucesso.", "sucesso")
    return redirect(url_for("painel_backup"))

@app.route("/restaurar_backup", methods=["GET", "POST"])
@login_obrigatorio
@admin_obrigatorio
def restaurar_backup():
    if request.method == "POST":
        arquivo = request.files.get("arquivo_backup")

        if not arquivo:
            flash("Selecione um arquivo de backup JSON.", "erro")
            return redirect(url_for("restaurar_backup"))

        try:
            backup = json.load(arquivo)

            dados = backup.get("dados", {})

            conn = conectar()
            cursor = conn.cursor()

            cursor.execute("DELETE FROM movimentacoes")
            cursor.execute("DELETE FROM produtos")
            cursor.execute("DELETE FROM categorias")

            for categoria in dados.get("categorias", []):
                cursor.execute("""
                    INSERT INTO categorias (id, nome, descricao, criado_em)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (id) DO NOTHING
                """, (
                    categoria.get("id"),
                    categoria.get("nome"),
                    categoria.get("descricao"),
                    categoria.get("criado_em")
                ))

            for produto in dados.get("produtos", []):
                cursor.execute("""
                    INSERT INTO produtos (
                        id, nome, categoria_id, lote, codigo_barras,
                        data_vencimento, data_abertura, validade_apos_aberto_dias,
                        quantidade_atual, estoque_padrao, limite_alerta,
                        observacoes, tipo_estoque, criado_em
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (id) DO NOTHING
                """, (
                    produto.get("id"),
                    produto.get("nome"),
                    produto.get("categoria_id"),
                    produto.get("lote"),
                    produto.get("codigo_barras"),
                    produto.get("data_vencimento"),
                    produto.get("data_abertura"),
                    produto.get("validade_apos_aberto_dias"),
                    produto.get("quantidade_atual"),
                    produto.get("estoque_padrao"),
                    produto.get("limite_alerta"),
                    produto.get("observacoes"),
                    produto.get("tipo_estoque", "almoxarifado"),
                    produto.get("criado_em")
                ))

            for mov in dados.get("movimentacoes", []):
                cursor.execute("""
                    INSERT INTO movimentacoes (
                        id, produto_id, produto_nome, tipo_movimentacao,
                        quantidade, estoque_anterior, estoque_atual,
                        observacao, tipo_estoque, estoque_origem,
                        estoque_destino, usuario_id, usuario_nome,
                        data_movimentacao
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (id) DO NOTHING
                """, (
                    mov.get("id"),
                    mov.get("produto_id"),
                    mov.get("produto_nome"),
                    mov.get("tipo_movimentacao"),
                    mov.get("quantidade"),
                    mov.get("estoque_anterior"),
                    mov.get("estoque_atual"),
                    mov.get("observacao"),
                    mov.get("tipo_estoque"),
                    mov.get("estoque_origem"),
                    mov.get("estoque_destino"),
                    mov.get("usuario_id"),
                    mov.get("usuario_nome"),
                    mov.get("data_movimentacao")
                ))

            cursor.execute("SELECT setval('categorias_id_seq', COALESCE((SELECT MAX(id) FROM categorias), 1))")
            cursor.execute("SELECT setval('produtos_id_seq', COALESCE((SELECT MAX(id) FROM produtos), 1))")
            cursor.execute("SELECT setval('movimentacoes_id_seq', COALESCE((SELECT MAX(id) FROM movimentacoes), 1))")

            conn.commit()
            conn.close()

            flash("Backup restaurado com sucesso.", "sucesso")
            return redirect(url_for("produtos"))

        except Exception as erro:
            flash(f"Erro ao restaurar backup: {erro}", "erro")
            return redirect(url_for("restaurar_backup"))

    return render_template("restaurar_backup.html")

@app.route("/")
def index():
    if session.get("logado"):
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    garantir_tabela_usuarios()

    if request.method == "POST":
        usuario_digitado = (
            request.form.get("usuario")
            or request.form.get("login")
            or request.form.get("username")
            or ""
        ).strip().lower()

        senha_digitada = (
            request.form.get("senha")
            or request.form.get("password")
            or ""
        )

        usuario = buscar_usuario_login(usuario_digitado)

        if usuario and check_password_hash(usuario["senha_hash"], senha_digitada):
            session["logado"] = True
            session["usuario_id"] = usuario["id"]
            session["usuario_nome"] = usuario["nome"]
            session["usuario_login"] = usuario["usuario"]
            session["perfil"] = usuario["perfil"]

            flash("Login realizado com sucesso.", "sucesso")

            if usuario["perfil"] == "admin":
                return redirect(url_for("dashboard"))
            return redirect(url_for("produtos"))

        flash("Usuário ou senha inválidos.", "erro")

    return render_template("login.html")


@app.route("/logout")
@login_obrigatorio
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/licenca_bloqueada")
@login_obrigatorio
def licenca_bloqueada():
    info = situacao_licenca()
    return render_template("licenca_bloqueada.html", info_licenca=info)


@app.route("/painel_master_licenca_8472")
@login_obrigatorio
@admin_obrigatorio
def licenciamento():
    info = situacao_licenca()
    return render_template("licenciamento.html", info=info)


@app.route("/licenciamento/renovar", methods=["POST"])
@login_obrigatorio
@admin_obrigatorio
def renovar_licenca():
    empresa = request.form.get("empresa", "").strip() or "Cliente Atual"
    plano = request.form.get("plano", "").strip() or "Premium Mensal"
    chave = request.form.get("chave_licenca", "").strip() or "LICENCA-LOCAL"
    data_vencimento = request.form.get("data_vencimento", "").strip()
    dias_carencia = int(request.form.get("dias_carencia") or 5)
    observacoes = request.form.get("observacoes", "").strip()

    if not data_vencimento:
        flash("Informe a data de vencimento da licença.", "erro")
        return redirect(url_for("licenciamento"))

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("SELECT id FROM licencas ORDER BY id LIMIT 1")
    licenca = cursor.fetchone()

    if licenca:
        cursor.execute("""
            UPDATE licencas
            SET empresa = %s,
                plano = %s,
                status = 'ativo',
                data_vencimento = %s,
                dias_carencia = %s,
                chave_licenca = %s,
                observacoes = %s,
                atualizado_em = NOW()
            WHERE id = %s
        """, (
            empresa,
            plano,
            data_vencimento,
            dias_carencia,
            chave,
            observacoes,
            licenca["id"]
        ))
    else:
        cursor.execute("""
            INSERT INTO licencas (
                empresa, plano, status, data_vencimento, dias_carencia,
                chave_licenca, observacoes
            ) VALUES (%s, %s, 'ativo', %s, %s, %s, %s)
        """, (
            empresa,
            plano,
            data_vencimento,
            dias_carencia,
            chave,
            observacoes
        ))

    conn.commit()
    conn.close()

    flash("Licença renovada com sucesso.", "sucesso")
    return redirect(url_for("licenciamento"))


@app.route("/licenciamento/bloquear", methods=["POST"])
@login_obrigatorio
@admin_obrigatorio
def bloquear_licenca():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("UPDATE licencas SET status = 'bloqueado', atualizado_em = NOW() WHERE id = (SELECT id FROM licencas ORDER BY id LIMIT 1)")
    conn.commit()
    conn.close()

    flash("Licença bloqueada.", "sucesso")
    return redirect(url_for("licenciamento"))


@app.route("/licenciamento/ativar", methods=["POST"])
@login_obrigatorio
@admin_obrigatorio
def ativar_licenca():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("UPDATE licencas SET status = 'ativo', atualizado_em = NOW() WHERE id = (SELECT id FROM licencas ORDER BY id LIMIT 1)")
    conn.commit()
    conn.close()

    flash("Licença ativada.", "sucesso")
    return redirect(url_for("licenciamento"))


@app.route("/alertas/whatsapp/enviar", methods=["POST"])
@login_obrigatorio
@admin_obrigatorio
def enviar_alerta_whatsapp_manual():
    sucesso, mensagem = executar_alerta_whatsapp_automatico(forcar=True)
    flash(mensagem, "sucesso" if sucesso else "erro")
    return redirect(url_for("config_alertas"))


@app.route("/dashboard")
@login_obrigatorio
@licenca_obrigatoria
@admin_obrigatorio
def dashboard():

   #try:
    #   executar_alerta_whatsapp_automatico(forcar=False)
    #xcept Exception:
    #   pass

    conn = conectar()
    cursor = conn.cursor()

    produtos = listar_produtos_resumo_estoque()

    cursor.execute("""
        SELECT *
        FROM movimentacoes
        WHERE data_movimentacao >= NOW() - INTERVAL '30 days'
        ORDER BY data_movimentacao DESC
    """)
    movimentacoes_30 = cursor.fetchall()

    cursor.execute("""
        SELECT *
        FROM movimentacoes
        WHERE tipo_movimentacao IN ('transferencia_saida', 'transferencia_entrada')
        ORDER BY data_movimentacao DESC
        LIMIT 6
    """)
    transferencias_recentes = cursor.fetchall()

    cursor.execute("""
        SELECT produto_nome, SUM(quantidade) AS total_saida
        FROM movimentacoes
        WHERE tipo_movimentacao = 'saida'
          AND data_movimentacao >= NOW() - INTERVAL '30 days'
        GROUP BY produto_nome
        ORDER BY total_saida DESC
        LIMIT 5
    """)
    consumo_top = cursor.fetchall()

    cursor.execute("""
        SELECT 
            COALESCE(SUM(CASE WHEN tipo_movimentacao = 'entrada' THEN quantidade ELSE 0 END), 0) AS entradas,
            COALESCE(SUM(CASE WHEN tipo_movimentacao = 'saida' THEN quantidade ELSE 0 END), 0) AS saidas
        FROM movimentacoes
        WHERE data_movimentacao >= NOW() - INTERVAL '30 days'
    """)
    resumo_mov = cursor.fetchone()

    conn.close()

    produtos_status = []
    contadores = {
        "total": len(produtos),
        "vencidos": 0,
        "proximos": 0,
        "vencidos_abertos": 0,
        "estoque_baixo": 0,
        "estoque_zerado": 0,
        "entradas_mes": resumo_mov["entradas"] or 0,
        "saidas_mes": resumo_mov["saidas"] or 0,
        "movimentacoes_mes": len(movimentacoes_30),
        "almox_baixo": 0,
        "farmacia_baixo": 0,
        "almox_vencendo": 0,
        "farmacia_vencendo": 0,
        "aguardando_compra": 0
    }

    estoque_ok = 0
    estoque_baixo = 0
    estoque_zerado = 0
    alertas_inteligentes = []
    previsoes = []

    consumo_por_nome = {item["produto_nome"]: item["total_saida"] or 0 for item in consumo_top}

    for produto in produtos:
        status = calcular_status(produto)
        produtos_status.append({"produto": produto, "status": status})
        texto = status["texto"].lower()

        if "vencido após abertura" in texto:
            contadores["vencidos_abertos"] += 1
            alertas_inteligentes.append(f"{produto['nome']} está vencido após abertura.")
        elif "vencido" in texto:
            contadores["vencidos"] += 1
            alertas_inteligentes.append(f"{produto['nome']} está vencido pela validade original.")

        if "vence" in texto:
            contadores["proximos"] += 1
            if produto["tipo_estoque"] == "farmacia_satelite":
                contadores["farmacia_vencendo"] += 1
            else:
                contadores["almox_vencendo"] += 1
            alertas_inteligentes.append(f"{produto['nome']} precisa de atenção: {status['texto']}.")

        if produto["quantidade_atual"] == 0:
            contadores["estoque_zerado"] += 1
            estoque_zerado += 1
            alertas_inteligentes.append(f"{produto['nome']} está com estoque zerado.")
        elif produto["quantidade_atual"] <= produto["limite_alerta"]:
            contadores["estoque_baixo"] += 1
            if produto["tipo_estoque"] == "farmacia_satelite":
                contadores["farmacia_baixo"] += 1
            else:
                contadores["almox_baixo"] += 1
            estoque_baixo += 1
            alertas_inteligentes.append(f"{produto['nome']} está abaixo do limite de estoque.")
        else:
            estoque_ok += 1

        if produto["quantidade_atual"] <= produto["limite_alerta"]:
            contadores["aguardando_compra"] += 1

        saidas_30 = consumo_por_nome.get(produto["nome"], 0)
        if saidas_30 > 0:
            media_diaria = float(saidas_30) / 30
            dias_estimados = int(produto["quantidade_atual"] / media_diaria) if produto["quantidade_atual"] > 0 else 0
            previsoes.append({
                "produto": produto["nome"],
                "dias": dias_estimados,
                "estoque": produto["quantidade_atual"],
                "media": round(media_diaria, 2)
            })

    grafico_movimentacoes = {
        "labels": ["Entradas", "Saídas"],
        "dados": [int(contadores["entradas_mes"]), int(contadores["saidas_mes"])]
    }
    grafico_estoque = {"labels": ["OK", "Baixo", "Zerado"], "dados": [estoque_ok, estoque_baixo, estoque_zerado]}
    grafico_consumo = {
        "labels": [item["produto_nome"] for item in consumo_top],
        "dados": [int(item["total_saida"] or 0) for item in consumo_top]
    }

    ranking_consumo = [{"produto": item["produto_nome"], "total": int(item["total_saida"] or 0)} for item in consumo_top]
    produtos_criticos = [item for item in produtos_status if item["status"]["tipo"] != "success"][:8]
    previsoes = sorted(previsoes, key=lambda x: x["dias"])[:5]
    alertas_inteligentes = alertas_inteligentes[:8]

    return render_template(
        "dashboard.html",
        contadores=contadores,
        produtos_status=produtos_status,
        produtos_criticos=produtos_criticos,
        ranking_consumo=ranking_consumo,
        previsoes=previsoes,
        alertas_inteligentes=alertas_inteligentes,
        grafico_movimentacoes=grafico_movimentacoes,
        grafico_estoque=grafico_estoque,
        grafico_consumo=grafico_consumo,
        info_licenca=situacao_licenca(),
        transferencias_recentes=transferencias_recentes
    )


def montar_central_categorias():
    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM categorias ORDER BY nome")
    categorias = cursor.fetchall()

    dados = []

    for categoria in categorias:
        produtos_cat = listar_produtos_resumo_estoque(categoria_id=categoria["id"])

        total_produtos = len(produtos_cat)
        estoque_total = sum(produto["quantidade_atual"] for produto in produtos_cat)
        estoque_baixo = 0
        vencendo = 0
        vencidos = 0

        for produto in produtos_cat:
            status = calcular_status(produto)
            texto = status["texto"].lower()

            if produto["quantidade_atual"] <= produto["limite_alerta"]:
                estoque_baixo += 1

            if "vencido" in texto:
                vencidos += 1
            elif "vence" in texto:
                vencendo += 1

        dados.append({
            "id": categoria["id"],
            "nome": categoria["nome"],
            "total_produtos": total_produtos,
            "estoque_total": estoque_total,
            "estoque_baixo": estoque_baixo,
            "vencendo": vencendo,
            "vencidos": vencidos
        })

    conn.close()
    return dados


@app.route("/categorias")
@login_obrigatorio
@licenca_obrigatoria
@estoque_obrigatorio
def categorias():
    categorias = montar_central_categorias()
    return render_template("categorias.html", categorias=categorias)


@app.route("/categorias/nova", methods=["GET", "POST"])
@login_obrigatorio
@alteracao_permitida
@estoque_obrigatorio
def nova_categoria():
    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        descricao = request.form.get("descricao", "").strip()

        if not nome:
            flash("O nome da categoria é obrigatório.", "erro")
            return render_template("nova_categoria.html")

        conn = conectar()
        cursor = conn.cursor()
        try:
            cursor.execute("INSERT INTO categorias (nome, descricao) VALUES (%s, %s)", (nome, descricao))
            conn.commit()
            flash("Categoria cadastrada com sucesso.", "sucesso")
            return redirect(url_for("categorias"))
        except psycopg2.IntegrityError:
            conn.rollback()
            flash("Essa categoria já existe.", "erro")
        finally:
            conn.close()

    return render_template("nova_categoria.html")


@app.route("/categorias/editar/<int:id>", methods=["GET", "POST"])
@login_obrigatorio
@alteracao_permitida
@admin_obrigatorio
def editar_categoria(id):
    conn = conectar()
    cursor = conn.cursor()

    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        descricao = request.form.get("descricao", "").strip()

        if not nome:
            flash("O nome da categoria é obrigatório.", "erro")
        else:
            try:
                cursor.execute("UPDATE categorias SET nome = %s, descricao = %s WHERE id = %s", (nome, descricao, id))
                conn.commit()
                flash("Categoria atualizada com sucesso.", "sucesso")
                conn.close()
                return redirect(url_for("categorias"))
            except psycopg2.IntegrityError:
                conn.rollback()
                flash("Essa categoria já existe.", "erro")

    cursor.execute("SELECT * FROM categorias WHERE id = %s", (id,))
    categoria = cursor.fetchone()
    conn.close()

    if not categoria:
        flash("Categoria não encontrada.", "erro")
        return redirect(url_for("categorias"))

    return render_template("editar_categoria.html", categoria=categoria)


@app.route("/categorias/excluir/<int:id>", methods=["POST"])
@login_obrigatorio
@alteracao_permitida
@admin_obrigatorio
def excluir_categoria(id):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) AS total FROM produtos WHERE categoria_id = %s", (id,))
    total = cursor.fetchone()["total"]

    if total > 0:
        flash("Não é possível excluir uma categoria com produtos vinculados.", "erro")
    else:
        cursor.execute("DELETE FROM categorias WHERE id = %s", (id,))
        conn.commit()
        flash("Categoria excluída com sucesso.", "sucesso")

    conn.close()
    return redirect(url_for("categorias"))


def listar_categorias():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM categorias ORDER BY nome")
    categorias = cursor.fetchall()
    conn.close()
    return categorias


def registrar_movimentacao(
    cursor,
    produto_id,
    produto_nome,
    tipo_movimentacao,
    quantidade,
    estoque_anterior,
    estoque_atual,
    observacao,
    tipo_estoque="almoxarifado",
    estoque_origem=None,
    estoque_destino=None
):
    usuario_id = session.get("usuario_id")
    usuario_nome = session.get("usuario_nome") or session.get("usuario_login") or "Sistema"

    cursor.execute("""
        INSERT INTO movimentacoes (
            produto_id, produto_nome, tipo_movimentacao, quantidade,
            estoque_anterior, estoque_atual, observacao, tipo_estoque,
            estoque_origem, estoque_destino, usuario_id, usuario_nome
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        produto_id,
        produto_nome,
        tipo_movimentacao,
        quantidade,
        estoque_anterior,
        estoque_atual,
        observacao,
        tipo_estoque,
        estoque_origem,
        estoque_destino,
        usuario_id,
        usuario_nome
    ))


def buscar_produtos_agrupados(tipo_estoque=None, busca="", categoria_id="", filtro="todos"):
    conn = conectar()
    cursor = conn.cursor()

    query = """
        SELECT
            pe.id AS produto_estoque_id,
            p.id AS produto_id,
            p.nome,
            p.codigo_barras,
            p.fabricante,
            p.unidade_medida,
            p.validade_apos_aberto_dias,
            p.observacoes,
            c.nome AS categoria_nome,
            pe.tipo_estoque,
            pe.estoque_padrao,
            pe.limite_alerta,
            COALESCE(SUM(l.quantidade_atual), 0) AS quantidade_atual,
            COUNT(l.id) FILTER (WHERE l.ativo = TRUE) AS total_lotes,
            MIN(l.data_vencimento) FILTER (WHERE l.ativo = TRUE AND l.quantidade_atual > 0) AS proximo_vencimento
        FROM produto_estoques pe
        JOIN produtos p ON p.id = pe.produto_id
        JOIN categorias c ON c.id = p.categoria_id
        LEFT JOIN lotes l ON l.produto_estoque_id = pe.id AND l.ativo = TRUE
        WHERE COALESCE(p.ativo, TRUE) = TRUE
          AND COALESCE(pe.ativo, TRUE) = TRUE
    """
    params = []

    if tipo_estoque:
        query += " AND pe.tipo_estoque = %s"
        params.append(tipo_estoque)

    if busca:
        query += " AND (p.nome ILIKE %s OR p.codigo_barras ILIKE %s)"
        params.extend([f"%{busca}%", f"%{busca}%"])

    if categoria_id:
        query += " AND p.categoria_id = %s"
        params.append(categoria_id)

    query += """
        GROUP BY pe.id, p.id, p.nome, p.codigo_barras, p.fabricante,
                 p.unidade_medida, p.validade_apos_aberto_dias, p.observacoes,
                 c.nome, pe.tipo_estoque, pe.estoque_padrao, pe.limite_alerta
        ORDER BY p.nome, pe.tipo_estoque
    """

    cursor.execute(query, params)
    produtos_lista = cursor.fetchall()
    conn.close()

    resultado = []
    for produto in produtos_lista:
        quantidade = produto["quantidade_atual"] or 0
        limite = produto["limite_alerta"] or 0
        status = {"tipo": "ok", "texto": "Produto OK"}

        if quantidade <= 0:
            status = {"tipo": "danger", "texto": "Estoque zerado"}
        elif quantidade <= limite:
            status = {"tipo": "warning", "texto": "Estoque baixo"}

        mostrar = True
        if filtro == "estoque":
            mostrar = quantidade > 0 and quantidade <= limite
        elif filtro == "zerado":
            mostrar = quantidade <= 0

        if mostrar:
            resultado.append({"produto": produto, "status": status, "vencido": False})

    return resultado


def buscar_lotes_produto_estoque(produto_estoque_id):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT
            l.*,
            pe.tipo_estoque,
            p.id AS produto_id,
            p.nome,
            p.codigo_barras,
            p.validade_apos_aberto_dias,
            c.nome AS categoria_nome
        FROM lotes l
        JOIN produto_estoques pe ON pe.id = l.produto_estoque_id
        JOIN produtos p ON p.id = pe.produto_id
        JOIN categorias c ON c.id = p.categoria_id
        WHERE l.produto_estoque_id = %s
          AND l.ativo = TRUE
        ORDER BY l.data_vencimento ASC, l.id ASC
    """, (produto_estoque_id,))
    lotes = cursor.fetchall()
    conn.close()
    return lotes


def buscar_lotes_vencidos_bloqueados(tipo_estoque=None, busca="", categoria_id=""):
    conn = conectar()
    cursor = conn.cursor()
    query = """
        SELECT
            l.*,
            pe.id AS produto_estoque_id,
            pe.tipo_estoque,
            p.id AS produto_id,
            p.nome,
            p.codigo_barras,
            c.nome AS categoria_nome
        FROM lotes l
        JOIN produto_estoques pe ON pe.id = l.produto_estoque_id
        JOIN produtos p ON p.id = pe.produto_id
        JOIN categorias c ON c.id = p.categoria_id
        WHERE l.ativo = TRUE
          AND l.quantidade_atual > 0
          AND l.data_vencimento < CURRENT_DATE
          AND COALESCE(p.ativo, TRUE) = TRUE
          AND COALESCE(pe.ativo, TRUE) = TRUE
    """
    params = []

    if tipo_estoque:
        query += " AND pe.tipo_estoque = %s"
        params.append(tipo_estoque)

    if busca:
        query += " AND (p.nome ILIKE %s OR p.codigo_barras ILIKE %s OR l.numero_lote ILIKE %s)"
        params.extend([f"%{busca}%", f"%{busca}%", f"%{busca}%"])

    if categoria_id:
        query += " AND p.categoria_id = %s"
        params.append(categoria_id)

    query += " ORDER BY l.data_vencimento ASC, p.nome"
    cursor.execute(query, params)
    lotes = cursor.fetchall()
    conn.close()
    return lotes


def listar_produtos_abertos_registrados(tipo_estoque="", busca="", status=""):
    conn = conectar()
    cursor = conn.cursor()
    query = """
        SELECT
            pa.*,
            c.nome AS categoria_nome,
            CASE
                WHEN pa.vencimento_apos_aberto < CURRENT_DATE THEN 'vencido'
                WHEN pa.vencimento_apos_aberto = CURRENT_DATE THEN 'vence_hoje'
                ELSE 'aberto'
            END AS status_abertura,
            (pa.vencimento_apos_aberto - CURRENT_DATE) AS dias_restantes
        FROM produtos_abertos pa
        LEFT JOIN produtos p ON p.id = pa.produto_id
        LEFT JOIN categorias c ON c.id = p.categoria_id
        WHERE 1=1
    """
    params = []

    if tipo_estoque:
        query += " AND pa.tipo_estoque = %s"
        params.append(tipo_estoque)

    if busca:
        query += " AND (pa.produto_nome ILIKE %s OR pa.numero_lote ILIKE %s)"
        params.extend([f"%{busca}%", f"%{busca}%"])

    if status == "vencido":
        query += " AND pa.vencimento_apos_aberto < CURRENT_DATE"
    elif status == "vence_hoje":
        query += " AND pa.vencimento_apos_aberto = CURRENT_DATE"
    elif status == "aberto":
        query += " AND pa.vencimento_apos_aberto > CURRENT_DATE"

    query += " ORDER BY pa.vencimento_apos_aberto ASC, pa.criado_em DESC"
    cursor.execute(query, params)
    registros = cursor.fetchall()
    conn.close()
    return registros


def listar_produtos_resumo_estoque(tipo_estoque=None, categoria_id=None, busca=None):
    conn = conectar()
    cursor = conn.cursor()

    query = """
        SELECT
            pe.id AS id,
            p.id AS produto_id,
            p.nome,
            p.categoria_id,
            p.codigo_barras,
            p.fabricante,
            p.unidade_medida,
            p.validade_apos_aberto_dias,
            p.observacoes,
            c.nome AS categoria_nome,
            pe.tipo_estoque,
            pe.estoque_padrao,
            pe.limite_alerta,
            COALESCE(SUM(l.quantidade_atual), 0) AS quantidade_atual,
            MIN(l.data_vencimento) FILTER (WHERE l.ativo = TRUE AND l.quantidade_atual > 0) AS data_vencimento,
            MAX(l.data_abertura) AS data_abertura,
            STRING_AGG(NULLIF(l.numero_lote, ''), ', ' ORDER BY l.data_vencimento) AS lote,
            COUNT(l.id) FILTER (WHERE l.ativo = TRUE) AS total_lotes
        FROM produto_estoques pe
        JOIN produtos p ON p.id = pe.produto_id
        JOIN categorias c ON c.id = p.categoria_id
        LEFT JOIN lotes l ON l.produto_estoque_id = pe.id AND l.ativo = TRUE
        WHERE COALESCE(p.ativo, TRUE) = TRUE
          AND COALESCE(pe.ativo, TRUE) = TRUE
    """
    params = []

    if tipo_estoque:
        query += " AND pe.tipo_estoque = %s"
        params.append(tipo_estoque)

    if categoria_id:
        query += " AND p.categoria_id = %s"
        params.append(categoria_id)

    if busca:
        query += " AND (p.nome ILIKE %s OR p.codigo_barras ILIKE %s)"
        params.extend([f"%{busca}%", f"%{busca}%"])

    query += """
        GROUP BY pe.id, p.id, p.nome, p.categoria_id, p.codigo_barras,
                 p.fabricante, p.unidade_medida, p.validade_apos_aberto_dias,
                 p.observacoes, c.nome, pe.tipo_estoque,
                 pe.estoque_padrao, pe.limite_alerta
        ORDER BY p.nome, pe.tipo_estoque
    """

    cursor.execute(query, params)
    produtos = cursor.fetchall()
    conn.close()
    return produtos


def registrar_movimentacao_lote(cursor, lote, tipo_movimentacao, quantidade, quantidade_anterior, quantidade_posterior, motivo, origem=None, destino=None):
    usuario_id = session.get("usuario_id")
    usuario_nome = session.get("usuario_nome") or session.get("usuario_login") or "Sistema"
    cursor.execute("""
        INSERT INTO movimentacoes (
            lote_id, produto_id, produto_nome, tipo_movimentacao,
            quantidade, estoque_anterior, estoque_atual,
            quantidade_anterior, quantidade_posterior,
            observacao, motivo, tipo_estoque, origem, destino,
            estoque_origem, estoque_destino, usuario_id, usuario_nome
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        lote["id"],
        lote["produto_id"],
        lote["nome"],
        tipo_movimentacao,
        quantidade,
        quantidade_anterior,
        quantidade_posterior,
        quantidade_anterior,
        quantidade_posterior,
        motivo,
        motivo,
        lote["tipo_estoque"],
        origem,
        destino,
        origem,
        destino,
        usuario_id,
        usuario_nome
    ))


def buscar_lote_para_movimentacao(cursor, lote_id):
    cursor.execute("""
        SELECT
            l.*,
            pe.tipo_estoque,
            pe.estoque_padrao,
            pe.limite_alerta,
            p.id AS produto_id,
            p.nome
        FROM lotes l
        JOIN produto_estoques pe ON pe.id = l.produto_estoque_id
        JOIN produtos p ON p.id = pe.produto_id
        WHERE l.id = %s
          AND l.ativo = TRUE
    """, (lote_id,))
    return cursor.fetchone()


def buscar_lote_por_codigo_scanner(cursor, codigo_barras, tipo_estoque):
    cursor.execute("""
        SELECT
            l.*,
            l.id AS lote_id,
            pe.id AS produto_estoque_id,
            pe.tipo_estoque,
            pe.estoque_padrao,
            pe.limite_alerta,
            p.id AS produto_id,
            p.nome,
            p.codigo_barras,
            p.unidade_medida,
            p.validade_apos_aberto_dias,
            c.nome AS categoria_nome
        FROM lotes l
        JOIN produto_estoques pe ON pe.id = l.produto_estoque_id
        JOIN produtos p ON p.id = pe.produto_id
        LEFT JOIN categorias c ON c.id = p.categoria_id
        WHERE p.codigo_barras = %s
          AND pe.tipo_estoque = %s
          AND p.ativo = TRUE
          AND pe.ativo = TRUE
          AND l.ativo = TRUE
          AND l.quantidade_atual > 0
          AND (l.data_vencimento IS NULL OR l.data_vencimento >= CURRENT_DATE)
        ORDER BY l.data_vencimento ASC NULLS LAST, l.id ASC
        LIMIT 1
    """, (codigo_barras, tipo_estoque))
    return cursor.fetchone()


def buscar_lote_operacional_por_produto_estoque(cursor, produto_estoque_id):
    cursor.execute("""
        SELECT
            l.*,
            l.id AS lote_id,
            pe.id AS produto_estoque_id,
            pe.tipo_estoque,
            pe.estoque_padrao,
            pe.limite_alerta,
            p.id AS produto_id,
            p.nome,
            p.codigo_barras,
            p.unidade_medida,
            p.validade_apos_aberto_dias,
            c.nome AS categoria_nome
        FROM lotes l
        JOIN produto_estoques pe ON pe.id = l.produto_estoque_id
        JOIN produtos p ON p.id = pe.produto_id
        LEFT JOIN categorias c ON c.id = p.categoria_id
        WHERE pe.id = %s
          AND p.ativo = TRUE
          AND pe.ativo = TRUE
          AND l.ativo = TRUE
          AND l.quantidade_atual > 0
          AND (l.data_vencimento IS NULL OR l.data_vencimento >= CURRENT_DATE)
        ORDER BY l.data_vencimento ASC NULLS LAST, l.id ASC
        LIMIT 1
    """, (produto_estoque_id,))
    return cursor.fetchone()


@app.route("/lotes/<int:lote_id>/retirar", methods=["POST"])
@login_obrigatorio
@alteracao_permitida
@estoque_obrigatorio
def retirar_lote(lote_id):
    try:
        quantidade = int(request.form.get("quantidade", 0))
    except ValueError:
        quantidade = 0

    motivo = (
        request.form.get("motivo", "").strip()
        or request.form.get("observacao", "").strip()
        or "Saída de lote"
    )

    if quantidade <= 0:
        flash("Quantidade inválida.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    conn = conectar()
    cursor = conn.cursor()
    lote = buscar_lote_para_movimentacao(cursor, lote_id)

    if not lote:
        conn.close()
        flash("Lote não encontrado.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    quantidade_anterior = lote["quantidade_atual"]
    if quantidade > quantidade_anterior:
        conn.close()
        flash("Não é possível retirar mais do que a quantidade atual do lote.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    quantidade_posterior = quantidade_anterior - quantidade
    cursor.execute("UPDATE lotes SET quantidade_atual = %s WHERE id = %s", (quantidade_posterior, lote_id))
    registrar_movimentacao_lote(
        cursor,
        lote,
        "saida",
        quantidade,
        quantidade_anterior,
        quantidade_posterior,
        motivo,
        lote["tipo_estoque"],
        None
    )
    conn.commit()
    conn.close()

    flash("Saída registrada no lote.", "sucesso")
    return redirect(request.referrer or url_for("produtos"))


@app.route("/lotes/<int:lote_id>/repor", methods=["POST"])
@login_obrigatorio
@alteracao_permitida
@estoque_obrigatorio
def repor_lote(lote_id):
    try:
        quantidade = int(request.form.get("quantidade", 0))
    except ValueError:
        quantidade = 0

    motivo = (
        request.form.get("motivo", "").strip()
        or request.form.get("observacao", "").strip()
        or "Reposição de lote"
    )

    if quantidade <= 0:
        flash("Quantidade inválida.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    conn = conectar()
    cursor = conn.cursor()
    lote = buscar_lote_para_movimentacao(cursor, lote_id)

    if not lote:
        conn.close()
        flash("Lote não encontrado.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    quantidade_anterior = lote["quantidade_atual"]
    quantidade_posterior = quantidade_anterior + quantidade
    cursor.execute("UPDATE lotes SET quantidade_atual = %s WHERE id = %s", (quantidade_posterior, lote_id))
    registrar_movimentacao_lote(
        cursor,
        lote,
        "reposicao",
        quantidade,
        quantidade_anterior,
        quantidade_posterior,
        motivo,
        None,
        lote["tipo_estoque"]
    )
    conn.commit()
    conn.close()

    flash("Reposição registrada no lote.", "sucesso")
    return redirect(request.referrer or url_for("produtos"))


@app.route("/lotes/<int:lote_id>/baixar_vencido", methods=["POST"])
@login_obrigatorio
@alteracao_permitida
@estoque_obrigatorio
def baixar_lote_vencido(lote_id):
    motivo = (
        request.form.get("motivo", "").strip()
        or request.form.get("observacao", "").strip()
        or "Baixa por vencimento"
    )

    conn = conectar()
    cursor = conn.cursor()
    lote = buscar_lote_para_movimentacao(cursor, lote_id)

    if not lote:
        conn.close()
        flash("Lote nao encontrado.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    data_vencimento = converter_data(lote["data_vencimento"])
    if not data_vencimento or data_vencimento >= date.today():
        conn.close()
        flash("Esse lote ainda nao esta vencido.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    quantidade_anterior = lote["quantidade_atual"]
    if quantidade_anterior <= 0:
        conn.close()
        flash("Esse lote ja esta sem saldo.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    cursor.execute("""
        UPDATE lotes
        SET quantidade_atual = 0,
            ativo = FALSE,
            observacoes = CONCAT(COALESCE(observacoes, ''), %s)
        WHERE id = %s
    """, (
        f"\nBaixado por vencimento em {datetime.now().strftime('%d/%m/%Y %H:%M')}. Motivo: {motivo}",
        lote_id
    ))
    registrar_movimentacao_lote(
        cursor,
        lote,
        "baixa_vencimento",
        quantidade_anterior,
        quantidade_anterior,
        0,
        motivo,
        lote["tipo_estoque"],
        None
    )

    conn.commit()
    conn.close()

    flash("Lote vencido baixado do estoque util e mantido no historico.", "sucesso")
    return redirect(request.referrer or url_for("produtos"))

@app.route("/lotes/<int:lote_id>/transferir", methods=["POST"])
@login_obrigatorio
@alteracao_permitida
@estoque_obrigatorio
def transferir_lote(lote_id):
    try:
        quantidade = int(request.form.get("quantidade", 0))
    except ValueError:
        quantidade = 0

    destino = request.form.get("destino", "").strip()
    motivo = (
        request.form.get("motivo", "").strip()
        or request.form.get("observacao", "").strip()
        or f"Transferência para {nome_tipo_estoque(destino)}"
    )

    if destino not in ESTOQUES:
        flash("Estoque de destino inválido.", "erro")
        return redirect(request.referrer or url_for("produtos"))
    if quantidade <= 0:
        flash("Quantidade inválida.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    conn = conectar()
    cursor = conn.cursor()
    lote = buscar_lote_para_movimentacao(cursor, lote_id)

    if not lote:
        conn.close()
        flash("Lote não encontrado.", "erro")
        return redirect(request.referrer or url_for("produtos"))
    if lote["tipo_estoque"] == destino:
        conn.close()
        flash("O destino deve ser diferente da origem.", "erro")
        return redirect(request.referrer or url_for("produtos"))
    if quantidade > lote["quantidade_atual"]:
        conn.close()
        flash("Não é possível transferir mais do que a quantidade atual do lote.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    origem_anterior = lote["quantidade_atual"]
    origem_posterior = origem_anterior - quantidade
    cursor.execute("UPDATE lotes SET quantidade_atual = %s WHERE id = %s", (origem_posterior, lote_id))

    cursor.execute("""
        INSERT INTO produto_estoques (
            produto_id, tipo_estoque, estoque_padrao, limite_alerta, ativo
        ) VALUES (%s, %s, %s, %s, TRUE)
        ON CONFLICT (produto_id, tipo_estoque)
        DO UPDATE SET ativo = TRUE
        RETURNING id
    """, (lote["produto_id"], destino, lote["estoque_padrao"], lote["limite_alerta"]))
    produto_estoque_destino_id = cursor.fetchone()["id"]

    cursor.execute("""
        SELECT *
        FROM lotes
        WHERE produto_estoque_id = %s
          AND COALESCE(numero_lote, '') = COALESCE(%s, '')
          AND data_vencimento = %s
          AND ativo = TRUE
        LIMIT 1
    """, (produto_estoque_destino_id, lote["numero_lote"], lote["data_vencimento"]))
    lote_destino = cursor.fetchone()

    if lote_destino:
        destino_anterior = lote_destino["quantidade_atual"]
        destino_posterior = destino_anterior + quantidade
        cursor.execute("UPDATE lotes SET quantidade_atual = %s WHERE id = %s", (destino_posterior, lote_destino["id"]))
        lote_destino_id = lote_destino["id"]
    else:
        destino_anterior = 0
        destino_posterior = quantidade
        cursor.execute("""
            INSERT INTO lotes (
                produto_estoque_id, numero_lote, data_vencimento, data_abertura,
                quantidade_inicial, quantidade_atual, observacoes, ativo
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, TRUE)
            RETURNING id
        """, (
            produto_estoque_destino_id,
            lote["numero_lote"],
            lote["data_vencimento"],
            lote["data_abertura"],
            quantidade,
            quantidade,
            lote["observacoes"]
        ))
        lote_destino_id = cursor.fetchone()["id"]

    registrar_movimentacao_lote(
        cursor,
        lote,
        "transferencia_saida",
        quantidade,
        origem_anterior,
        origem_posterior,
        motivo,
        lote["tipo_estoque"],
        destino
    )

    lote_destino_mov = dict(lote)
    lote_destino_mov["id"] = lote_destino_id
    lote_destino_mov["tipo_estoque"] = destino
    registrar_movimentacao_lote(
        cursor,
        lote_destino_mov,
        "transferencia_entrada",
        quantidade,
        destino_anterior,
        destino_posterior,
        motivo,
        lote["tipo_estoque"],
        destino
    )

    conn.commit()
    conn.close()

    flash("Transferência de lote realizada com rastreabilidade.", "sucesso")
    return redirect(request.referrer or url_for("produtos"))


@app.route("/produto_estoque/<int:produto_estoque_id>/retirar", methods=["POST"])
@login_obrigatorio
@alteracao_permitida
@estoque_obrigatorio
def retirar_produto_estoque(produto_estoque_id):
    conn = conectar()
    cursor = conn.cursor()
    lote = buscar_lote_operacional_por_produto_estoque(cursor, produto_estoque_id)
    conn.close()

    if not lote:
        flash("Nenhum lote valido disponivel para retirada. Verifique os lotes ou faca nova entrada.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    return retirar_lote(lote["id"])


@app.route("/produto_estoque/<int:produto_estoque_id>/devolver", methods=["POST"])
@login_obrigatorio
@alteracao_permitida
@estoque_obrigatorio
def devolver_produto_estoque(produto_estoque_id):
    try:
        quantidade = int(request.form.get("quantidade", 0))
    except ValueError:
        quantidade = 0

    motivo = (
        request.form.get("motivo", "").strip()
        or request.form.get("observacao", "").strip()
        or "Devolucao de produto"
    )

    if quantidade <= 0:
        flash("Quantidade invalida.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    conn = conectar()
    cursor = conn.cursor()
    lote = buscar_lote_operacional_por_produto_estoque(cursor, produto_estoque_id)

    if not lote:
        conn.close()
        flash("Nenhum lote valido disponivel para devolucao. Use Nova entrada se for lote novo.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    quantidade_anterior = lote["quantidade_atual"]
    quantidade_posterior = quantidade_anterior + quantidade
    cursor.execute("UPDATE lotes SET quantidade_atual = %s WHERE id = %s", (quantidade_posterior, lote["id"]))
    registrar_movimentacao_lote(
        cursor,
        lote,
        "devolucao",
        quantidade,
        quantidade_anterior,
        quantidade_posterior,
        motivo,
        None,
        lote["tipo_estoque"]
    )
    conn.commit()
    conn.close()

    flash("Devolucao registrada no lote.", "sucesso")
    return redirect(request.referrer or url_for("produtos"))


@app.route("/produto_estoque/<int:produto_estoque_id>/transferir", methods=["POST"])
@login_obrigatorio
@alteracao_permitida
@estoque_obrigatorio
def transferir_produto_estoque(produto_estoque_id):
    destino = request.form.get("destino", "").strip()

    conn = conectar()
    cursor = conn.cursor()
    lote = buscar_lote_operacional_por_produto_estoque(cursor, produto_estoque_id)
    conn.close()

    if not lote:
        flash("Nenhum lote valido disponivel para transferencia. Verifique os lotes ou faca nova entrada.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    if lote["tipo_estoque"] == "carrinho_urgencia":
        flash("O Carrinho de Urgencia nao deve enviar itens para repor outro estoque.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    if destino == "almoxarifado":
        flash("O Almoxarifado e estoque de origem. Nao envie material de volta para ele por transferencia operacional.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    return transferir_lote(lote["id"])

@app.route("/produtos")
@app.route("/produtos/<tipo_estoque>")
@login_obrigatorio
@licenca_obrigatoria
@estoque_obrigatorio
def produtos(tipo_estoque=None):

    #try:
    #   executar_alerta_whatsapp_automatico(forcar=False)
    #except Exception:
    #    pass

    filtro = request.args.get("filtro", "todos")
    busca = request.args.get("busca", "").strip()
    categoria_id = request.args.get("categoria_id", "").strip()

    if tipo_estoque not in [None, "almoxarifado", "farmacia", "farmacia_satelite", "carrinho_urgencia"]:
        tipo_estoque = None

    tipo_banco = None
    if tipo_estoque == "almoxarifado":
        tipo_banco = "almoxarifado"
    elif tipo_estoque in ["farmacia", "farmacia_satelite"]:
        tipo_banco = "farmacia_satelite"
    elif tipo_estoque == "carrinho_urgencia":
        tipo_banco = "carrinho_urgencia"

    categorias_lista = listar_categorias()
    todos_produtos_status = buscar_produtos_agrupados(tipo_banco, busca, categoria_id, filtro)
    mostrar_alertas_gerais = tipo_banco is None and filtro == "todos" and not busca and not categoria_id
    produtos_vencidos = buscar_lotes_vencidos_bloqueados() if mostrar_alertas_gerais else []

    pagina = int(request.args.get("pagina", 1))
    por_pagina = 5
    total_produtos = len(todos_produtos_status)
    total_paginas = (total_produtos + por_pagina - 1) // por_pagina

    inicio = (pagina - 1) * por_pagina
    fim = inicio + por_pagina

    produtos_status = todos_produtos_status[inicio:fim]

    return render_template(
        "produtos.html",
        produtos_status=produtos_status,
        produtos_vencidos=produtos_vencidos,
        filtro=filtro,
        busca=busca,
        categoria_id=categoria_id,
        categorias=categorias_lista,
        tipo_estoque=tipo_banco,
        tipo_estoque_nome=nome_tipo_estoque(tipo_banco) if tipo_banco else "Todos os Estoques",
        pagina=pagina,
        total_paginas=total_paginas,
        total_resultados=total_produtos,
        mostrar_alertas_gerais=mostrar_alertas_gerais,
    )


@app.route("/produtos/nova_entrada", methods=["GET", "POST"])
@login_obrigatorio
@alteracao_permitida
@estoque_obrigatorio
def nova_entrada_produto():
    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT produtos.*, categorias.nome AS categoria_nome
        FROM produtos
        JOIN categorias ON categorias.id = produtos.categoria_id
        WHERE COALESCE(produtos.ativo, TRUE) = TRUE
        ORDER BY produtos.nome
    """)
    produtos_lista = cursor.fetchall()
    cursor.execute("""
        SELECT produto_id, tipo_estoque, estoque_padrao, limite_alerta
        FROM produto_estoques
        WHERE COALESCE(ativo, TRUE) = TRUE
    """)
    produto_configuracoes = {}
    for config in cursor.fetchall():
        produto_configuracoes.setdefault(str(config["produto_id"]), {})[config["tipo_estoque"]] = {
            "estoque_padrao": config["estoque_padrao"],
            "limite_alerta": config["limite_alerta"],
        }

    if request.method == "POST":
        produto_id = request.form.get("produto_id")
        tipo_estoque = request.form.get("tipo_estoque", "almoxarifado").strip()
        numero_lote = request.form.get("numero_lote", "").strip()
        data_vencimento = request.form.get("data_vencimento", "").strip()
        tipo_unidade_entrada = request.form.get("tipo_unidade_entrada", "unidade").strip()
        quantidade_embalagens = int(request.form.get("quantidade_embalagens") or 0)
        unidades_por_embalagem = int(request.form.get("unidades_por_embalagem") or 0)
        quantidade = quantidade_embalagens * unidades_por_embalagem
        estoque_padrao = int(request.form.get("estoque_padrao") or 0)
        limite_alerta = int(request.form.get("limite_alerta") or 0)
        observacoes = request.form.get("observacoes", "").strip()

        erros = []
        if not produto_id:
            erros.append("Selecione o produto.")
        if tipo_estoque not in ESTOQUES:
            erros.append("Estoque inválido.")
        if not data_vencimento:
            erros.append("Informe a validade do lote.")
        if quantidade <= 0:
            erros.append("Quantidade recebida deve ser maior que zero.")
        if quantidade_embalagens <= 0:
            erros.append("Quantidade de caixas/pacotes/unidades deve ser maior que zero.")
        if unidades_por_embalagem <= 0:
            erros.append("Quantidade por caixa/pacote deve ser maior que zero.")
        if estoque_padrao < 0:
            erros.append("Estoque padrão não pode ser negativo.")
        if limite_alerta < 0:
            erros.append("Limite de alerta não pode ser negativo.")

        if erros:
            for erro in erros:
                flash(erro, "erro")
            conn.close()
            return render_template(
                "nova_entrada.html",
                produtos=produtos_lista,
                produto_configuracoes=produto_configuracoes,
                form=request.form
            )

        try:
            detalhe_quantidade = (
                f"Entrada em {tipo_unidade_entrada}: "
                f"{quantidade_embalagens} x {unidades_por_embalagem} = {quantidade} unidade(s)."
            )
            observacoes_lote = f"{observacoes}\n{detalhe_quantidade}".strip()

            cursor.execute("""
                INSERT INTO produto_estoques (
                    produto_id, tipo_estoque, estoque_padrao, limite_alerta, ativo
                ) VALUES (%s, %s, %s, %s, TRUE)
                ON CONFLICT (produto_id, tipo_estoque)
                DO UPDATE SET
                    estoque_padrao = EXCLUDED.estoque_padrao,
                    limite_alerta = EXCLUDED.limite_alerta,
                    ativo = TRUE
                RETURNING id
            """, (produto_id, tipo_estoque, estoque_padrao, limite_alerta))
            produto_estoque_id = cursor.fetchone()["id"]

            cursor.execute("""
                INSERT INTO lotes (
                    produto_estoque_id, numero_lote, data_vencimento,
                    quantidade_inicial, quantidade_atual, observacoes, ativo
                ) VALUES (%s, %s, %s, %s, %s, %s, TRUE)
                RETURNING id
            """, (
                produto_estoque_id,
                numero_lote,
                data_vencimento,
                quantidade,
                quantidade,
                observacoes_lote
            ))
            lote_id = cursor.fetchone()["id"]

            cursor.execute("""
                SELECT
                    l.id, l.quantidade_atual, pe.tipo_estoque,
                    p.id AS produto_id, p.nome
                FROM lotes l
                JOIN produto_estoques pe ON pe.id = l.produto_estoque_id
                JOIN produtos p ON p.id = pe.produto_id
                WHERE l.id = %s
            """, (lote_id,))
            lote = cursor.fetchone()
            registrar_movimentacao_lote(
                cursor,
                lote,
                "entrada",
                quantidade,
                0,
                quantidade,
                "Entrada de novo lote",
                None,
                tipo_estoque
            )

            conn.commit()
            flash("Entrada cadastrada e lote criado com sucesso.", "sucesso")
            return redirect(url_for("produtos", tipo_estoque=tipo_estoque))
        except Exception as erro:
            conn.rollback()
            flash(f"Erro ao cadastrar entrada: {erro}", "erro")

    conn.close()
    return render_template(
        "nova_entrada.html",
        produtos=produtos_lista,
        produto_configuracoes=produto_configuracoes,
        form=request.form if request.method == "POST" else None
    )


@app.route("/produtos/<int:produto_estoque_id>/lotes")
@login_obrigatorio
@licenca_obrigatoria
@estoque_obrigatorio
def lotes_produto(produto_estoque_id):
    lotes = buscar_lotes_produto_estoque(produto_estoque_id)
    if not lotes:
        flash("Nenhum lote encontrado para esse produto no estoque.", "erro")
        return redirect(url_for("produtos"))

    return render_template("lotes_produto.html", lotes=lotes, produto=lotes[0])


@app.route("/produtos/aberto", methods=["GET", "POST"])
@login_obrigatorio
@alteracao_permitida
@estoque_obrigatorio
def produto_aberto():
    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT produtos.*, categorias.nome AS categoria_nome
        FROM produtos
        JOIN categorias ON categorias.id = produtos.categoria_id
        WHERE COALESCE(produtos.ativo, TRUE) = TRUE
        ORDER BY produtos.nome
    """)
    produtos_lista = cursor.fetchall()

    if request.method == "POST":
        lote_id = request.form.get("lote_id")
        data_abertura = request.form.get("data_abertura", "").strip()
        validade_apos_aberto_dias = request.form.get("validade_apos_aberto_dias", "").strip()
        observacoes = request.form.get("observacoes", "").strip()

        erros = []
        abertura = converter_data(data_abertura)
        if not lote_id:
            erros.append("Selecione o lote que será aberto.")
        if not abertura:
            erros.append("Informe a data de abertura.")
        elif abertura > date.today():
            erros.append("Data de abertura não pode ser futura.")
        validade_dias = None
        if validade_apos_aberto_dias:
            try:
                validade_dias = int(validade_apos_aberto_dias)
                if validade_dias <= 0:
                    erros.append("Validade apos aberto deve ser maior que zero.")
            except (TypeError, ValueError):
                erros.append("Informe a validade apos aberto em dias ou deixe em branco para usar a validade do lote.")

        if erros:
            for erro in erros:
                flash(erro, "erro")
            conn.close()
            return redirect(url_for("produto_aberto"))

        lote = buscar_lote_para_movimentacao(cursor, lote_id)
        if not lote:
            conn.close()
            flash("Lote não encontrado.", "erro")
            return redirect(url_for("produto_aberto"))
        if lote["quantidade_atual"] <= 0:
            conn.close()
            flash("Lote sem quantidade disponível para abertura.", "erro")
            return redirect(url_for("produto_aberto"))

        quantidade_anterior = lote["quantidade_atual"]
        quantidade_posterior = quantidade_anterior - 1
        data_vencimento_lote = converter_data(lote["data_vencimento"])

        if validade_dias:
            vencimento_apos_aberto = abertura + timedelta(days=validade_dias)
            origem_validade = f"validade especifica apos aberto: {validade_dias} dia(s)"
        else:
            vencimento_apos_aberto = data_vencimento_lote
            validade_dias = max(0, (vencimento_apos_aberto - abertura).days)
            origem_validade = "sem validade especifica apos aberto; calculado pela validade original do lote"

        if vencimento_apos_aberto < abertura:
            conn.close()
            flash("A validade do lote e anterior a data de abertura.", "erro")
            return redirect(url_for("produto_aberto"))

        motivo = (
            f"Produto aberto em {formatar_data(abertura)}. "
            f"Vence em {formatar_data(vencimento_apos_aberto)} ({origem_validade})."
        )
        if observacoes:
            motivo = f"{motivo} {observacoes}"

        cursor.execute("UPDATE lotes SET quantidade_atual = %s, data_abertura = COALESCE(data_abertura, %s) WHERE id = %s", (
            quantidade_posterior,
            data_abertura,
            lote_id
        ))
        cursor.execute("""
            INSERT INTO produtos_abertos (
                lote_id, produto_id, produto_nome, tipo_estoque, numero_lote,
                data_abertura, validade_apos_aberto_dias, vencimento_apos_aberto,
                quantidade, usuario_id, usuario_nome, observacoes
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 1, %s, %s, %s)
        """, (
            lote_id,
            lote["produto_id"],
            lote["nome"],
            lote["tipo_estoque"],
            lote["numero_lote"],
            data_abertura,
            validade_dias,
            vencimento_apos_aberto,
            session.get("usuario_id"),
            session.get("usuario_nome") or session.get("usuario_login") or "Sistema",
            observacoes
        ))
        registrar_movimentacao_lote(
            cursor,
            lote,
            "retirada",
            1,
            quantidade_anterior,
            quantidade_posterior,
            motivo,
            lote["tipo_estoque"],
            None
        )

        conn.commit()
        conn.close()
        flash("Produto aberto registrado e 1 unidade retirada do lote.", "sucesso")
        return redirect(url_for("produtos", tipo_estoque=lote["tipo_estoque"]))

    produto_id = request.args.get("produto_id", "").strip()
    codigo_barras = request.args.get("codigo_barras", "").strip()
    tipo_estoque = request.args.get("tipo_estoque", "almoxarifado").strip()
    if tipo_estoque not in ESTOQUES:
        tipo_estoque = "almoxarifado"

    lotes = []
    if produto_id or codigo_barras:
        query = """
            SELECT
                l.*,
                pe.tipo_estoque,
                p.id AS produto_id,
                p.nome,
                p.codigo_barras,
                c.nome AS categoria_nome
            FROM lotes l
            JOIN produto_estoques pe ON pe.id = l.produto_estoque_id
            JOIN produtos p ON p.id = pe.produto_id
            JOIN categorias c ON c.id = p.categoria_id
            WHERE pe.tipo_estoque = %s
              AND l.ativo = TRUE
              AND l.quantidade_atual > 0
        """
        params = [tipo_estoque]
        if codigo_barras:
            query += " AND p.codigo_barras = %s"
            params.append(codigo_barras)
        elif produto_id:
            query += " AND p.id = %s"
            params.append(produto_id)
        query += " ORDER BY l.data_vencimento ASC, l.id ASC"
        cursor.execute(query, params)
        lotes = cursor.fetchall()
        if not lotes:
            flash("Nenhum lote disponível encontrado para esse produto nesse estoque.", "erro")

    conn.close()
    return render_template(
        "produto_aberto.html",
        produtos=produtos_lista,
        produto_id=produto_id,
        codigo_barras=codigo_barras,
        tipo_estoque=tipo_estoque,
        lotes=lotes,
        hoje=date.today().isoformat()
    )


@app.route("/produtos/abertos")
@login_obrigatorio
@licenca_obrigatoria
@estoque_obrigatorio
def produtos_abertos():
    tipo_estoque = request.args.get("tipo_estoque", "").strip()
    busca = request.args.get("busca", "").strip()
    status = request.args.get("status", "").strip()

    if tipo_estoque not in ["", "almoxarifado", "farmacia_satelite", "carrinho_urgencia"]:
        tipo_estoque = ""
    if status not in ["", "aberto", "vence_hoje", "vencido"]:
        status = ""

    registros = listar_produtos_abertos_registrados(tipo_estoque, busca, status)
    total_vencidos = len([item for item in registros if item["status_abertura"] == "vencido"])
    total_vence_hoje = len([item for item in registros if item["status_abertura"] == "vence_hoje"])

    return render_template(
        "produtos_abertos.html",
        registros=registros,
        tipo_estoque=tipo_estoque,
        busca=busca,
        status=status,
        total_vencidos=total_vencidos,
        total_vence_hoje=total_vence_hoje,
    )


@app.route("/produtos/baixar_estoque/<int:id>", methods=["POST"])
@login_obrigatorio
@alteracao_permitida
@estoque_obrigatorio
def baixar_estoque(id):
    try:
        quantidade = int(request.form.get("quantidade", 0))
    except ValueError:
        quantidade = 0

    observacao = request.form.get("observacao", "").strip()

    if quantidade <= 0:
        flash("Quantidade inválida.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM produtos WHERE id = %s", (id,))
    produto = cursor.fetchone()

    if not produto:
        conn.close()
        flash("Produto não encontrado.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    if produto_esta_vencido(produto):
        conn.close()
        flash("Produto vencido bloqueado para uso. Faça descarte ou ajuste administrativo.", "erro")
        return redirect(request.referrer or url_for("produtos"))


    estoque_anterior = produto["quantidade_atual"]

    if quantidade > estoque_anterior:
        conn.close()
        flash("Não é possível baixar mais do que o estoque atual.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    estoque_atual = estoque_anterior - quantidade

    cursor.execute("UPDATE produtos SET quantidade_atual = %s WHERE id = %s", (estoque_atual, id))
    registrar_movimentacao(cursor, id, produto["nome"], "retirada", quantidade, estoque_anterior, estoque_atual, observacao, produto["tipo_estoque"])

    conn.commit()
    conn.close()

    flash("Estoque baixado e histórico registrado com sucesso.", "sucesso")
    return redirect(request.referrer or url_for("produtos"))


@app.route("/produtos/repor_estoque/<int:id>", methods=["POST"])
@login_obrigatorio
@alteracao_permitida
@estoque_obrigatorio
def repor_estoque(id):
    try:
        quantidade = int(request.form.get("quantidade", 0))
    except ValueError:
        quantidade = 0

    observacao = request.form.get("observacao", "").strip()

    if quantidade <= 0:
        flash("Quantidade inválida.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM produtos WHERE id = %s", (id,))
    produto = cursor.fetchone()

    if not produto:
        conn.close()
        flash("Produto não encontrado.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    estoque_anterior = produto["quantidade_atual"]
    estoque_atual = estoque_anterior + quantidade

    cursor.execute("UPDATE produtos SET quantidade_atual = %s WHERE id = %s", (estoque_atual, id))
    registrar_movimentacao(cursor, id, produto["nome"], "reposicao", quantidade, estoque_anterior, estoque_atual, observacao, produto["tipo_estoque"])

    conn.commit()
    conn.close()

    flash("Estoque reposto e histórico registrado com sucesso.", "sucesso")
    return redirect(request.referrer or url_for("produtos"))


def validar_produto(form):
    erros = []
    nome = form.get("nome", "").strip()
    categoria_id = form.get("categoria_id", "").strip()

    if not nome:
        erros.append("Nome do produto é obrigatório.")
    if not categoria_id:
        erros.append("Categoria é obrigatória.")

    return erros


@app.route("/produtos/devolver_estoque/<int:id>", methods=["POST"])
@login_obrigatorio
@alteracao_permitida
@estoque_obrigatorio
def devolver_estoque(id):
    try:
        quantidade = int(request.form.get("quantidade", 0))
    except ValueError:
        quantidade = 0

    observacao = request.form.get("observacao", "").strip()

    if quantidade <= 0:
        flash("Quantidade inválida.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM produtos WHERE id = %s", (id,))
    produto = cursor.fetchone()

    if not produto:
        conn.close()
        flash("Produto não encontrado.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    estoque_anterior = produto["quantidade_atual"]
    estoque_atual = estoque_anterior + quantidade

    cursor.execute("UPDATE produtos SET quantidade_atual = %s WHERE id = %s", (estoque_atual, id))
    registrar_movimentacao(cursor, id, produto["nome"], "devolucao", quantidade, estoque_anterior, estoque_atual, observacao, produto["tipo_estoque"])

    conn.commit()
    conn.close()

    flash("Produto devolvido ao estoque e histórico registrado com sucesso.", "sucesso")
    return redirect(request.referrer or url_for("produtos"))



@app.route("/produtos/transferir/<int:id>", methods=["POST"])
@login_obrigatorio
@alteracao_permitida
@estoque_obrigatorio
def transferir_estoque(id):
    try:
        quantidade = int(request.form.get("quantidade", 0))
    except ValueError:
        quantidade = 0

    observacao = request.form.get("observacao", "").strip()
    destino = request.form.get("destino", "").strip()

    if destino not in ["almoxarifado", "farmacia_satelite"]:
        flash("Estoque de destino inválido.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    if quantidade <= 0:
        flash("Quantidade inválida.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM produtos WHERE id = %s", (id,))
    origem = cursor.fetchone()

    if not origem:
        conn.close()
        flash("Produto de origem não encontrado.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    if produto_esta_vencido(origem):
        conn.close()
        flash("Produto vencido bloqueado para transferência. Faça descarte ou ajuste administrativo.", "erro")
        return redirect(request.referrer or url_for("produtos"))


    if origem["tipo_estoque"] == destino:
        conn.close()
        flash("O estoque de destino deve ser diferente do estoque de origem.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    if quantidade > origem["quantidade_atual"]:
        conn.close()
        flash("Não é possível transferir mais do que o estoque atual.", "erro")
        return redirect(request.referrer or url_for("produtos"))

    cursor.execute("""
        SELECT *
        FROM produtos
        WHERE nome = %s
          AND COALESCE(lote, '') = COALESCE(%s, '')
          AND tipo_estoque = %s
        LIMIT 1
    """, (
        origem["nome"],
        origem["lote"],
        destino
    ))

    produto_destino = cursor.fetchone()

    estoque_origem_anterior = origem["quantidade_atual"]
    estoque_origem_atual = estoque_origem_anterior - quantidade

    cursor.execute(
        "UPDATE produtos SET quantidade_atual = %s WHERE id = %s",
        (estoque_origem_atual, origem["id"])
    )

    if produto_destino:
        estoque_destino_anterior = produto_destino["quantidade_atual"]
        estoque_destino_atual = estoque_destino_anterior + quantidade

        cursor.execute(
            "UPDATE produtos SET quantidade_atual = %s WHERE id = %s",
            (estoque_destino_atual, produto_destino["id"])
        )
        produto_destino_id = produto_destino["id"]
    else:
        estoque_destino_anterior = 0
        estoque_destino_atual = quantidade

        cursor.execute("""
            INSERT INTO produtos (
                nome,
                categoria_id,
                lote,
                codigo_barras,
                data_vencimento,
                data_abertura,
                validade_apos_aberto_dias,
                quantidade_atual,
                estoque_padrao,
                limite_alerta,
                observacoes,
                tipo_estoque
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            origem["nome"],
            origem["categoria_id"],
            origem["lote"],
            origem.get("codigo_barras"),
            origem["data_vencimento"],
            origem["data_abertura"],
            origem["validade_apos_aberto_dias"],
            quantidade,
            origem["estoque_padrao"],
            origem["limite_alerta"],
            origem["observacoes"],
            destino
        ))

        produto_destino_id = cursor.fetchone()["id"]

    descricao = observacao or f"Transferência para {nome_tipo_estoque(destino)}"

    registrar_movimentacao(
        cursor,
        origem["id"],
        origem["nome"],
        "transferencia_saida",
        quantidade,
        estoque_origem_anterior,
        estoque_origem_atual,
        descricao,
        origem["tipo_estoque"],
        origem["tipo_estoque"],
        destino
    )

    registrar_movimentacao(
        cursor,
        produto_destino_id,
        origem["nome"],
        "transferencia_entrada",
        quantidade,
        estoque_destino_anterior,
        estoque_destino_atual,
        descricao,
        destino,
        origem["tipo_estoque"],
        destino
    )

    conn.commit()
    conn.close()

    flash("Transferência realizada com sucesso.", "sucesso")
    return redirect(request.referrer or url_for("produtos"))


@app.route("/produtos/novo", methods=["GET", "POST"])
@login_obrigatorio
@alteracao_permitida
@estoque_obrigatorio
def novo_produto():
    categorias_lista = listar_categorias()

    if request.method == "POST":
        erros = validar_produto(request.form)

        if erros:
            for erro in erros:
                flash(erro, "erro")
            return render_template("novo_produto.html", categorias=categorias_lista, form=request.form)

        nome = request.form.get("nome", "").strip()
        categoria_id = request.form.get("categoria_id")
        codigo_barras = request.form.get("codigo_barras", "").strip()
        unidade_medida = request.form.get("unidade_medida", "").strip()
        observacoes = request.form.get("observacoes", "").strip()
        tipo_estoque = request.form.get("tipo_estoque", "almoxarifado").strip()
        numero_lote = request.form.get("numero_lote", "").strip()
        data_vencimento = request.form.get("data_vencimento", "").strip()
        tipo_unidade_entrada = request.form.get("tipo_unidade_entrada", unidade_medida or "unidade").strip()
        quantidade_embalagens = int(request.form.get("quantidade_embalagens") or 0)
        unidades_por_embalagem = int(request.form.get("unidades_por_embalagem") or 0)
        quantidade = quantidade_embalagens * unidades_por_embalagem
        estoque_padrao = int(request.form.get("estoque_padrao") or 0)
        limite_alerta = int(request.form.get("limite_alerta") or 0)

        if tipo_estoque not in ESTOQUES:
            erros.append("Estoque inválido.")
        if not data_vencimento:
            erros.append("Informe a validade do lote.")
        if quantidade <= 0:
            erros.append("Quantidade recebida deve ser maior que zero.")
        if quantidade_embalagens <= 0:
            erros.append("Quantidade de caixas/pacotes/unidades deve ser maior que zero.")
        if unidades_por_embalagem <= 0:
            erros.append("Quantidade por caixa/pacote deve ser maior que zero.")
        if estoque_padrao < 0:
            erros.append("Estoque padrão não pode ser negativo.")
        if limite_alerta < 0:
            erros.append("Limite de alerta não pode ser negativo.")

        if erros:
            for erro in erros:
                flash(erro, "erro")
            return render_template("novo_produto.html", categorias=categorias_lista, form=request.form)

        conn = conectar()
        cursor = conn.cursor()

        try:
            cursor.execute("""
                SELECT id
                FROM produtos
                WHERE LOWER(TRIM(nome)) = LOWER(TRIM(%s))
                  AND categoria_id = %s
                LIMIT 1
            """, (nome, categoria_id))

            produto_existente = cursor.fetchone()

            if produto_existente:
                flash("Esse produto já existe. Use Nova entrada para cadastrar um lote.", "erro")
                return render_template("novo_produto.html", categorias=categorias_lista, form=request.form)

            cursor.execute("""
                INSERT INTO produtos (
                    nome, categoria_id, codigo_barras,
                    unidade_medida,
                    observacoes, ativo
                ) VALUES (%s, %s, %s, %s, %s, TRUE)
                RETURNING id
            """, (
                nome,
                categoria_id,
                codigo_barras,
                unidade_medida,
                observacoes
            ))
            produto_id = cursor.fetchone()["id"]

            cursor.execute("""
                INSERT INTO produto_estoques (
                    produto_id, tipo_estoque, estoque_padrao, limite_alerta, ativo
                ) VALUES (%s, %s, %s, %s, TRUE)
                RETURNING id
            """, (produto_id, tipo_estoque, estoque_padrao, limite_alerta))
            produto_estoque_id = cursor.fetchone()["id"]

            detalhe_quantidade = (
                f"Entrada inicial em {tipo_unidade_entrada}: "
                f"{quantidade_embalagens} x {unidades_por_embalagem} = {quantidade} unidade(s)."
            )
            observacoes_lote = f"{observacoes}\n{detalhe_quantidade}".strip()

            cursor.execute("""
                INSERT INTO lotes (
                    produto_estoque_id, numero_lote, data_vencimento,
                    quantidade_inicial, quantidade_atual, observacoes, ativo
                ) VALUES (%s, %s, %s, %s, %s, %s, TRUE)
                RETURNING id
            """, (
                produto_estoque_id,
                numero_lote,
                data_vencimento,
                quantidade,
                quantidade,
                observacoes_lote
            ))
            lote_id = cursor.fetchone()["id"]

            cursor.execute("""
                SELECT
                    l.id, l.quantidade_atual, pe.tipo_estoque,
                    p.id AS produto_id, p.nome
                FROM lotes l
                JOIN produto_estoques pe ON pe.id = l.produto_estoque_id
                JOIN produtos p ON p.id = pe.produto_id
                WHERE l.id = %s
            """, (lote_id,))
            lote = cursor.fetchone()
            registrar_movimentacao_lote(
                cursor,
                lote,
                "entrada",
                quantidade,
                0,
                quantidade,
                "Entrada inicial do produto",
                None,
                tipo_estoque
            )

            conn.commit()
            flash("Produto cadastrado com lote e entrada inicial.", "sucesso")
            return redirect(url_for("produtos", tipo_estoque=tipo_estoque))

        except Exception as erro:
            conn.rollback()
            flash(f"Erro ao cadastrar produto: {erro}", "erro")
            return render_template("novo_produto.html", categorias=categorias_lista, form=request.form)

        finally:
            conn.close()

    return render_template("novo_produto.html", categorias=categorias_lista, form=None)


@app.route("/produtos/editar/<int:id>", methods=["GET", "POST"])
@login_obrigatorio
@alteracao_permitida
@estoque_obrigatorio
def editar_produto(id):
    categorias_lista = listar_categorias()
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM produtos WHERE id = %s", (id,))
    produto = cursor.fetchone()

    if not produto:
        conn.close()
        flash("Produto não encontrado.", "erro")
        return redirect(url_for("produtos"))

    if request.method == "POST":
        erros = validar_produto(request.form)

        if erros:
            for erro in erros:
                flash(erro, "erro")
            conn.close()
            return render_template("editar_produto.html", produto=produto, categorias=categorias_lista)

        cursor.execute("""
            UPDATE produtos SET
                nome = %s,
                categoria_id = %s,
                codigo_barras = %s,
                unidade_medida = %s,
                observacoes = %s
            WHERE id = %s
        """, (
            request.form.get("nome").strip(),
            request.form.get("categoria_id"),
            request.form.get("codigo_barras", "").strip(),
            request.form.get("unidade_medida", "").strip(),
            request.form.get("observacoes", "").strip(),
            id
        ))
        conn.commit()
        conn.close()
        flash("Produto atualizado com sucesso.", "sucesso")
        return redirect(url_for("produtos"))

    conn.close()
    return render_template("editar_produto.html", produto=produto, categorias=categorias_lista)


@app.route("/produtos/excluir/<int:id>", methods=["POST"])
@login_obrigatorio
@alteracao_permitida
@estoque_obrigatorio
def excluir_produto(id):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM produtos WHERE id = %s", (id,))
    conn.commit()
    conn.close()
    flash("Produto excluído com sucesso.", "sucesso")
    return redirect(url_for("produtos"))


@app.route("/auditoria_usuarios")
@login_obrigatorio
@admin_obrigatorio
def auditoria_usuarios():
    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            COALESCE(usuario_nome, 'Sistema') AS usuario_nome,
            COUNT(*) AS total_movimentacoes,
            COALESCE(SUM(CASE WHEN tipo_movimentacao IN ('retirada', 'saida', 'transferencia_saida') THEN quantidade ELSE 0 END), 0) AS total_saidas,
            COALESCE(SUM(CASE WHEN tipo_movimentacao IN ('reposicao', 'entrada', 'devolucao', 'transferencia_entrada') THEN quantidade ELSE 0 END), 0) AS total_entradas,
            MAX(data_movimentacao) AS ultima_acao
        FROM movimentacoes
        GROUP BY COALESCE(usuario_nome, 'Sistema')
        ORDER BY total_movimentacoes DESC
    """)
    resumo = cursor.fetchall()

    cursor.execute("""
        SELECT *
        FROM movimentacoes
        ORDER BY data_movimentacao DESC
        LIMIT 30
    """)
    movimentacoes = cursor.fetchall()

    conn.close()

    return render_template("auditoria_usuarios.html", resumo=resumo, movimentacoes=movimentacoes)


@app.route("/historico")
@login_obrigatorio
@licenca_obrigatoria
@estoque_obrigatorio
def historico():
    busca = request.args.get("busca", "").strip()
    tipo = request.args.get("tipo", "").strip()
    tipo_estoque = request.args.get("tipo_estoque", "").strip()
    usuario_nome = request.args.get("usuario_nome", "").strip()

    conn = conectar()
    cursor = conn.cursor()

    query = "SELECT * FROM movimentacoes WHERE 1=1"
    params = []

    if busca:
        query += " AND produto_nome ILIKE %s"
        params.append(f"%{busca}%")

    if tipo:
        query += " AND tipo_movimentacao = %s"
        params.append(tipo)

    if tipo_estoque:
        query += " AND tipo_estoque = %s"
        params.append(tipo_estoque)

    if usuario_nome:
        query += " AND usuario_nome ILIKE %s"
        params.append(f"%{usuario_nome}%")

    query += " ORDER BY data_movimentacao DESC"

    cursor.execute(query, params)
    movimentacoes = cursor.fetchall()
    conn.close()

    return render_template(
        "historico.html",
        movimentacoes=movimentacoes,
        busca=busca,
        tipo=tipo,
        tipo_estoque=tipo_estoque,
        usuario_nome=usuario_nome,
        produto_nome=None
    )


@app.route("/historico/produto/<int:id>")
@login_obrigatorio
@licenca_obrigatoria
def historico_produto(id):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT nome FROM produtos WHERE id = %s", (id,))
    produto = cursor.fetchone()
    cursor.execute("SELECT * FROM movimentacoes WHERE produto_id = %s ORDER BY data_movimentacao DESC", (id,))
    movimentacoes = cursor.fetchall()
    conn.close()

    produto_nome = produto["nome"] if produto else "Produto removido"
    return render_template("historico.html", movimentacoes=movimentacoes, busca="", tipo="", produto_nome=produto_nome)


def montar_painel_relatorios():
    conn = conectar()
    cursor = conn.cursor()

    produtos = listar_produtos_resumo_estoque()

    cursor.execute("""
        SELECT *
        FROM movimentacoes
        ORDER BY data_movimentacao DESC
        LIMIT 12
    """)
    movimentacoes = cursor.fetchall()

    total_produtos = len(produtos)
    estoque_baixo = 0
    vencidos = 0
    proximos_vencimento = 0
    criticos = []
    vencendo = []

    hoje = date.today()

    for produto in produtos:
        status = calcular_status(produto)
        texto = status["texto"].lower()

        if produto["quantidade_atual"] <= produto["limite_alerta"]:
            estoque_baixo += 1
            criticos.append({
                "nome": produto["nome"],
                "categoria": produto["categoria_nome"],
                "quantidade": produto["quantidade_atual"],
                "tipo": "Estoque baixo",
                "status": "stock"
            })

        if "vencido" in texto:
            vencidos += 1
            criticos.append({
                "nome": produto["nome"],
                "categoria": produto["categoria_nome"],
                "quantidade": produto["quantidade_atual"],
                "tipo": "Vencido",
                "status": "danger"
            })

        elif "vence" in texto:
            proximos_vencimento += 1
            data_venc = converter_data(produto["data_vencimento"])
            dias = (data_venc - hoje).days if data_venc else "-"
            vencendo.append({
                "nome": produto["nome"],
                "data": formatar_data(produto["data_vencimento"]),
                "dias": dias
            })
            criticos.append({
                "nome": produto["nome"],
                "categoria": produto["categoria_nome"],
                "quantidade": produto["quantidade_atual"],
                "tipo": "Próximo do vencimento",
                "status": "warning"
            })

    total_entradas = 0
    total_saidas = 0
    reposicoes = 0
    transferencias = 0
    devolucoes = 0
    retiradas = 0

    for mov in movimentacoes:
        tipo = mov["tipo_movimentacao"]

        if tipo in ["reposicao", "entrada", "transferencia_entrada"]:
            total_entradas += mov["quantidade"]

        if tipo in ["retirada", "saida", "transferencia_saida"]:
            total_saidas += mov["quantidade"]

        if tipo == "reposicao" or tipo == "entrada":
            reposicoes += 1

        if tipo in ["transferencia_saida", "transferencia_entrada"]:
            transferencias += 1

        if tipo == "devolucao":
            devolucoes += 1

        if tipo == "retirada" or tipo == "saida":
            retiradas += 1

    cursor.execute("""
        SELECT produto_nome, SUM(quantidade) AS total
        FROM movimentacoes
        WHERE tipo_movimentacao IN ('retirada', 'saida')
        GROUP BY produto_nome
        ORDER BY total DESC
        LIMIT 6
    """)
    consumo = cursor.fetchall()

    top_consumo = [
        {
            "nome": item["produto_nome"],
            "total": item["total"]
        }
        for item in consumo
    ]

    movs_formatadas = []
    for mov in movimentacoes:
        movs_formatadas.append({
            "data": formatar_data(mov["data_movimentacao"]),
            "produto": mov["produto_nome"],
            "tipo": mov["tipo_movimentacao"].replace("_", " ").title(),
            "quantidade": mov["quantidade"],
            "estoque": nome_tipo_estoque(mov["tipo_estoque"]) if mov["tipo_estoque"] else "-"
        })

    conn.close()

    return {
        "total_produtos": total_produtos,
        "estoque_baixo": estoque_baixo,
        "vencidos": vencidos,
        "proximos_vencimento": proximos_vencimento,
        "total_entradas": total_entradas,
        "total_saidas": total_saidas,
        "reposicoes": reposicoes,
        "transferencias": transferencias,
        "devolucoes": devolucoes,
        "retiradas": retiradas,
        "criticos": criticos[:12],
        "top_consumo": top_consumo,
        "vencendo": vencendo[:8],
        "movimentacoes": movs_formatadas
    }


@app.route("/relatorios")
@login_obrigatorio
@licenca_obrigatoria
@admin_obrigatorio
def relatorios():
    relatorio = montar_painel_relatorios()
    return render_template("relatorios.html", relatorio=relatorio)


def nome_arquivo_relatorio(tipo_relatorio, formato, tipo_estoque=""):
    data = datetime.now().strftime("%Y-%m-%d")
    nome = tipo_relatorio.replace("_", "-")

    if tipo_estoque:
        nome += "_" + tipo_estoque.replace("_", "-")

    return f"{nome}_{data}.{formato}"


def aplicar_estilo_excel(ws, titulo, total_colunas):
    ws.insert_rows(1, 3)
    ws["A1"] = "Controle Oftalmo"
    ws["A2"] = titulo
    ws["A1"].font = Font(bold=True, size=16, color="0F172A")
    ws["A2"].font = Font(bold=True, size=13, color="2563EB")

    header_row = 4
    fill = PatternFill("solid", fgColor="0F172A")
    font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    for col in range(1, total_colunas + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for col in range(1, total_colunas + 1):
        letter = get_column_letter(col)
        maior = 12
        for cell in ws[letter]:
            if cell.value:
                maior = max(maior, len(str(cell.value)) + 2)
        ws.column_dimensions[letter].width = min(maior, 35)


def gerar_excel(titulo, cabecalhos, linhas, nome_arquivo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"
    ws.append(cabecalhos)

    for linha in linhas:
        ws.append(linha)

    aplicar_estilo_excel(ws, titulo, len(cabecalhos))

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    resposta = send_file(
        output,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        max_age=0
    )
    resposta.headers["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    resposta.headers["X-Content-Type-Options"] = "nosniff"
    resposta.headers["Cache-Control"] = "no-store"
    return resposta


def gerar_pdf(titulo, cabecalhos, linhas, nome_arquivo):
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=1 * cm, leftMargin=1 * cm, topMargin=1 * cm, bottomMargin=1 * cm)
    styles = getSampleStyleSheet()
    elementos = []

    logo_path = os.path.join(app.root_path, "static", "logo.png")
    if os.path.exists(logo_path):
        try:
            elementos.append(Image(logo_path, width=2.0 * cm, height=2.0 * cm))
        except Exception:
            pass

    elementos.append(Paragraph("<b>Controle Oftalmo</b>", styles["Title"]))
    elementos.append(Paragraph(titulo, styles["Heading2"]))
    elementos.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
    elementos.append(Spacer(1, 0.4 * cm))

    if not linhas:
        elementos.append(Paragraph("Nenhum dado encontrado para este relatório.", styles["Normal"]))
    else:
        dados = [cabecalhos]
        for linha in linhas:
            dados.append([str(valor) if valor is not None else "-" for valor in linha])

        tabela = Table(dados, repeatRows=1)
        tabela.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        elementos.append(tabela)

    doc.build(elementos)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=nome_arquivo, mimetype="application/pdf")


def buscar_produtos_para_relatorio(busca="", tipo_estoque=""):
    return listar_produtos_resumo_estoque(tipo_estoque=tipo_estoque or None, busca=busca or None)


def montar_relatorio_produtos(tipo_relatorio, busca="", tipo_estoque=""):
    produtos_lista = buscar_produtos_para_relatorio(busca, tipo_estoque)

    cabecalhos = ["Produto", "Categoria", "Local", "Código de Barras", "Lote", "Vencimento", "Aberto em", "Vence após aberto", "Estoque", "Estoque padrão", "Limite alerta", "Status"]
    linhas = []

    for produto in produtos_lista:
        status = calcular_status(produto)
        texto = status["texto"].lower()
        incluir = True

        if tipo_relatorio == "vencimentos":
            incluir = "vencido" in texto or "vence" in texto
        elif tipo_relatorio == "estoque":
            incluir = "estoque" in texto

        if incluir:
            linhas.append([
                produto["nome"],
                produto["categoria_nome"],
                nome_tipo_estoque(produto["tipo_estoque"]),
                produto.get("codigo_barras") or "-",
                produto["lote"] or "-",
                formatar_data(produto["data_vencimento"]),
                formatar_data(produto["data_abertura"]),
                formatar_data(status["vencimento_apos_aberto"]),
                produto["quantidade_atual"],
                produto["estoque_padrao"],
                produto["limite_alerta"],
                status["texto"]
            ])

    if tipo_relatorio == "produtos":
        titulo = "Relatório geral de produtos"
    elif tipo_relatorio == "vencimentos":
        titulo = "Relatório de vencidos e próximos do vencimento"
    else:
        titulo = "Relatório de estoque baixo e zerado"

    if tipo_estoque:
        titulo += f" - {nome_tipo_estoque(tipo_estoque)}"
    else:
        titulo += " - Todos os estoques"

    return titulo, cabecalhos, linhas


def montar_relatorio_movimentacoes(busca="", tipo_movimentacao="", tipo_estoque=""):
    conn = conectar()
    cursor = conn.cursor()
    query = "SELECT * FROM movimentacoes WHERE 1=1"
    params = []

    if busca:
        query += " AND produto_nome ILIKE %s"
        params.append(f"%{busca}%")

    if tipo_movimentacao:
        query += " AND tipo_movimentacao = %s"
        params.append(tipo_movimentacao)

    if tipo_estoque:
        query += " AND tipo_estoque = %s"
        params.append(tipo_estoque)

    query += " ORDER BY data_movimentacao DESC"
    cursor.execute(query, params)
    movimentacoes = cursor.fetchall()
    conn.close()

    cabecalhos = ["Data", "Produto", "Usuário", "Local", "Tipo", "Quantidade", "Estoque anterior", "Estoque atual", "Observação"]
    linhas = []

    for mov in movimentacoes:
        linhas.append([
            formatar_data(mov["data_movimentacao"]),
            mov["produto_nome"],
            mov.get("usuario_nome") or "Sistema",
            nome_tipo_estoque(mov["tipo_estoque"]),
            mov["tipo_movimentacao"].capitalize(),
            mov["quantidade"],
            mov["estoque_anterior"],
            mov["estoque_atual"],
            mov["observacao"] or "-"
        ])

    titulo = "Relatório de movimentações de estoque"
    if tipo_estoque:
        titulo += f" - {nome_tipo_estoque(tipo_estoque)}"
    else:
        titulo += " - Todos os estoques"

    return titulo, cabecalhos, linhas


@app.route("/relatorios/exportar")
@login_obrigatorio
@licenca_obrigatoria
@admin_obrigatorio
def exportar_relatorio():
    tipo_relatorio = request.args.get("tipo_relatorio", "produtos")
    formato = request.args.get("formato", "xlsx")
    busca = request.args.get("busca", "").strip()
    tipo_movimentacao = request.args.get("tipo_movimentacao", "").strip()
    tipo_estoque = request.args.get("tipo_estoque", "").strip()

    if tipo_estoque not in ["", "almoxarifado", "farmacia_satelite"]:
        tipo_estoque = ""

    if tipo_relatorio == "movimentacoes":
        titulo, cabecalhos, linhas = montar_relatorio_movimentacoes(busca, tipo_movimentacao, tipo_estoque)
    else:
        titulo, cabecalhos, linhas = montar_relatorio_produtos(tipo_relatorio, busca, tipo_estoque)

    if formato not in ["xlsx", "pdf"]:
        flash("Formato de relatório inválido.", "erro")
        return redirect(url_for("relatorios"))

    nome_arquivo = nome_arquivo_relatorio(tipo_relatorio, formato, tipo_estoque)

    if formato == "xlsx":
        return gerar_excel(titulo, cabecalhos, linhas, nome_arquivo)

    return gerar_pdf(titulo, cabecalhos, linhas, nome_arquivo)


def origens_permitidas_para(destino):
    if destino == "farmacia_satelite":
        return ["almoxarifado"]
    if destino == "carrinho_urgencia":
        return ["farmacia_satelite", "almoxarifado"]
    return []


def montar_saldos_por_produto(produtos):
    saldos = {}
    for produto in produtos:
        saldos.setdefault(produto["produto_id"], {})[produto["tipo_estoque"]] = produto
    return saldos


def buscar_sugestao_transferencia(produto, saldos_por_produto):
    necessario = max(0, produto["estoque_padrao"] - (produto["quantidade_valida"] or 0))
    if necessario <= 0:
        return None

    for origem in origens_permitidas_para(produto["tipo_estoque"]):
        origem_produto = saldos_por_produto.get(produto["produto_id"], {}).get(origem)
        if not origem_produto:
            continue

        sobra = max(0, (origem_produto["quantidade_valida"] or 0) - (origem_produto["limite_alerta"] or 0))
        if sobra <= 0:
            continue

        quantidade_transferir = min(necessario, sobra)
        if quantidade_transferir <= 0:
            continue

        return {
            "tipo": "reposicao",
            "origem": origem,
            "origem_nome": nome_tipo_estoque(origem),
            "destino": produto["tipo_estoque"],
            "destino_nome": nome_tipo_estoque(produto["tipo_estoque"]),
            "quantidade": quantidade_transferir,
            "estoque_disponivel": origem_produto["quantidade_valida"],
            "mensagem": f"Transferir {quantidade_transferir} de {nome_tipo_estoque(origem)} para {nome_tipo_estoque(produto['tipo_estoque'])}."
        }

    return None


def buscar_sugestao_vencimento(produto, saldos_por_produto):
    quantidade_proxima = produto["quantidade_proxima_vencimento"] or 0
    if quantidade_proxima <= 0:
        return None

    estoque = produto["tipo_estoque"]
    prazo = dias_alerta_vencimento(estoque)

    if estoque == "almoxarifado":
        return {
            "tipo": "rodizio_vencimento",
            "origem": "almoxarifado",
            "origem_nome": nome_tipo_estoque("almoxarifado"),
            "destino": "farmacia_satelite",
            "destino_nome": nome_tipo_estoque("farmacia_satelite"),
            "quantidade": quantidade_proxima,
            "mensagem": f"Lote vence em ate {prazo} dias: transferir {quantidade_proxima} do Almoxarifado para a Farmacia Satelite para uso primeiro."
        }

    if estoque == "farmacia_satelite":
        origem_produto = saldos_por_produto.get(produto["produto_id"], {}).get("almoxarifado")
        if origem_produto and (origem_produto["quantidade_valida"] or 0) > (origem_produto["limite_alerta"] or 0):
            return {
                "tipo": "troca_vencimento",
                "origem": "almoxarifado",
                "origem_nome": nome_tipo_estoque("almoxarifado"),
                "destino": "farmacia_satelite",
                "destino_nome": nome_tipo_estoque("farmacia_satelite"),
                "quantidade": min(quantidade_proxima, (origem_produto["quantidade_valida"] or 0) - (origem_produto["limite_alerta"] or 0)),
                "mensagem": f"Lote da Farmacia vence em ate {prazo} dias: usar primeiro e repor/trocar com lote valido do Almoxarifado."
            }
        return {
            "tipo": "vencimento_sem_transferencia",
            "origem": None,
            "origem_nome": "-",
            "destino": estoque,
            "destino_nome": nome_tipo_estoque(estoque),
            "quantidade": quantidade_proxima,
            "mensagem": f"Lote da Farmacia vence em ate {prazo} dias: priorizar uso e acompanhar baixa."
        }

    if estoque == "carrinho_urgencia":
        for origem in ["farmacia_satelite", "almoxarifado"]:
            origem_produto = saldos_por_produto.get(produto["produto_id"], {}).get(origem)
            if origem_produto and (origem_produto["quantidade_valida"] or 0) > 0:
                return {
                    "tipo": "troca_carrinho",
                    "origem": origem,
                    "origem_nome": nome_tipo_estoque(origem),
                    "destino": "carrinho_urgencia",
                    "destino_nome": nome_tipo_estoque("carrinho_urgencia"),
                    "quantidade": quantidade_proxima,
                    "mensagem": f"Carrinho vence em ate {prazo} dias: nao retirar para repor outro local; substituir por lote valido de {nome_tipo_estoque(origem)} e priorizar uso/baixa."
                }
        return {
            "tipo": "vencimento_carrinho",
            "origem": None,
            "origem_nome": "-",
            "destino": "carrinho_urgencia",
            "destino_nome": nome_tipo_estoque("carrinho_urgencia"),
            "quantidade": quantidade_proxima,
            "mensagem": f"Carrinho vence em ate {prazo} dias: nao sai para repor outro local; acompanhar uso, troca ou descarte."
        }

    return None


def buscar_itens_ordem_compra(tipo_estoque=""):
    conn = conectar()
    cursor = conn.cursor()
    query = """
        SELECT
            pe.id AS id,
            p.id AS produto_id,
            p.nome,
            c.nome AS categoria_nome,
            pe.tipo_estoque,
            pe.estoque_padrao,
            pe.limite_alerta,
            COALESCE(SUM(l.quantidade_atual), 0) AS quantidade_atual,
            COALESCE(SUM(l.quantidade_atual) FILTER (WHERE l.data_vencimento < CURRENT_DATE), 0) AS quantidade_vencida,
            COALESCE(SUM(l.quantidade_atual) FILTER (WHERE l.data_vencimento >= CURRENT_DATE), 0) AS quantidade_valida,
            COALESCE(SUM(l.quantidade_atual) FILTER (
                WHERE l.data_vencimento >= CURRENT_DATE
                  AND l.data_vencimento <= CURRENT_DATE + (
                      CASE
                          WHEN pe.tipo_estoque = 'carrinho_urgencia' THEN INTERVAL '30 days'
                          ELSE INTERVAL '15 days'
                      END
                  )
            ), 0) AS quantidade_proxima_vencimento,
            STRING_AGG(NULLIF(l.numero_lote, ''), ', ' ORDER BY l.data_vencimento) AS lote,
            STRING_AGG(NULLIF(l.numero_lote, ''), ', ' ORDER BY l.data_vencimento)
                FILTER (WHERE l.data_vencimento < CURRENT_DATE AND l.quantidade_atual > 0) AS lotes_vencidos,
            STRING_AGG(NULLIF(l.numero_lote, ''), ', ' ORDER BY l.data_vencimento)
                FILTER (
                    WHERE l.data_vencimento >= CURRENT_DATE
                      AND l.data_vencimento <= CURRENT_DATE + (
                          CASE
                              WHEN pe.tipo_estoque = 'carrinho_urgencia' THEN INTERVAL '30 days'
                              ELSE INTERVAL '15 days'
                          END
                      )
                      AND l.quantidade_atual > 0
                ) AS lotes_proximos_vencimento
        FROM produto_estoques pe
        JOIN produtos p ON p.id = pe.produto_id
        JOIN categorias c ON c.id = p.categoria_id
        LEFT JOIN lotes l ON l.produto_estoque_id = pe.id AND l.ativo = TRUE
        WHERE COALESCE(p.ativo, TRUE) = TRUE
          AND COALESCE(pe.ativo, TRUE) = TRUE
    """
    params = []

    query += """
        GROUP BY pe.id, p.id, p.nome, c.nome, pe.tipo_estoque, pe.estoque_padrao, pe.limite_alerta
        ORDER BY pe.tipo_estoque, p.nome
    """
    cursor.execute(query, params)
    produtos = cursor.fetchall()
    conn.close()

    saldos_por_produto = montar_saldos_por_produto(produtos)
    itens = []

    for produto in produtos:
        if tipo_estoque and produto["tipo_estoque"] != tipo_estoque:
            continue

        quantidade_vencida = produto["quantidade_vencida"] or 0
        quantidade_valida = produto["quantidade_valida"] or 0
        quantidade_proxima = produto["quantidade_proxima_vencimento"] or 0
        prazo_vencimento = dias_alerta_vencimento(produto["tipo_estoque"])
        tem_vencido = quantidade_vencida > 0
        tem_proximo_vencimento = quantidade_proxima > 0
        abaixo_limite = quantidade_valida <= produto["limite_alerta"]
        necessidade = max(0, produto["estoque_padrao"] - quantidade_valida)

        if not (tem_vencido or tem_proximo_vencimento or abaixo_limite):
            continue

        sugestao_transferencia = None
        sugestao_vencimento = None
        sugestao = 0

        if tem_vencido:
            sugestao = produto["estoque_padrao"]
            nivel = "critico"
            status = "Vencido"
            acao_sugerida = "Comprar"
            motivo_compra = "Lote vencido nao conta como estoque util. Comprar o padrao completo e baixar/descartar o vencido."
        else:
            if abaixo_limite:
                sugestao_transferencia = buscar_sugestao_transferencia(produto, saldos_por_produto)
                transferivel = sugestao_transferencia["quantidade"] if sugestao_transferencia else 0
                sugestao = max(0, necessidade - transferivel)
            if tem_proximo_vencimento:
                sugestao_vencimento = buscar_sugestao_vencimento(produto, saldos_por_produto)

            if quantidade_valida == 0:
                nivel = "critico"
                status = "Critico"
            elif tem_proximo_vencimento:
                nivel = "atencao"
                status = f"Vence em {prazo_vencimento} dias"
            else:
                nivel = "atencao"
                status = "Atencao"

            if sugestao_transferencia and sugestao > 0:
                acao_sugerida = "Transferir e comprar"
                motivo_compra = f"Estoque valido abaixo do limite. {sugestao_transferencia['mensagem']} Comprar o restante: {sugestao}."
            elif sugestao_transferencia:
                acao_sugerida = "Transferir"
                motivo_compra = f"Estoque valido abaixo do limite. {sugestao_transferencia['mensagem']}"
            elif sugestao_vencimento and not abaixo_limite:
                acao_sugerida = "Rodizio por vencimento"
                motivo_compra = sugestao_vencimento["mensagem"]
            elif sugestao_vencimento:
                acao_sugerida = "Comprar e rodiziar"
                motivo_compra = f"Estoque valido abaixo do limite e ha lote perto de vencer. {sugestao_vencimento['mensagem']}"
            else:
                acao_sugerida = "Comprar"
                motivo_compra = "Estoque valido abaixo do limite. Comprar ate completar o estoque padrao."

        acao_operacional = sugestao_transferencia or sugestao_vencimento

        itens.append({
            "id": produto["id"],
            "nome": produto["nome"],
            "categoria": produto["categoria_nome"],
            "lote": produto["lote"] or "-",
            "tipo_estoque": produto["tipo_estoque"],
            "tipo_estoque_nome": nome_tipo_estoque(produto["tipo_estoque"]),
            "quantidade_atual": produto["quantidade_atual"],
            "quantidade_valida": quantidade_valida,
            "quantidade_vencida": quantidade_vencida,
            "quantidade_proxima_vencimento": quantidade_proxima,
            "prazo_vencimento": prazo_vencimento,
            "lotes_vencidos": produto["lotes_vencidos"] or "-",
            "lotes_proximos_vencimento": produto["lotes_proximos_vencimento"] or "-",
            "limite_alerta": produto["limite_alerta"],
            "estoque_padrao": produto["estoque_padrao"],
            "sugestao": sugestao,
            "sugestao_transferencia": sugestao_transferencia,
            "sugestao_vencimento": sugestao_vencimento,
            "acao_operacional": acao_operacional,
            "acao_sugerida": acao_sugerida,
            "nivel": nivel,
            "status": status,
            "motivo_compra": motivo_compra
        })

    return itens

def separar_itens_por_estoque(itens):
    return {
        "almoxarifado": [item for item in itens if item["tipo_estoque"] == "almoxarifado"],
        "farmacia_satelite": [item for item in itens if item["tipo_estoque"] == "farmacia_satelite"],
        "carrinho_urgencia": [item for item in itens if item["tipo_estoque"] == "carrinho_urgencia"]
    }


@app.route("/ordem_compra")
@login_obrigatorio
@licenca_obrigatoria
@estoque_obrigatorio
def ordem_compra():
    tipo_estoque = request.args.get("tipo_estoque", "").strip()

    if tipo_estoque not in ["", "almoxarifado", "farmacia_satelite", "carrinho_urgencia"]:
        tipo_estoque = ""

    itens = buscar_itens_ordem_compra(tipo_estoque)
    separados = separar_itens_por_estoque(itens)

    total_sugerido = sum(item["sugestao"] for item in itens)
    total_critico = len([item for item in itens if item["nivel"] == "critico"])

    return render_template(
        "ordem_compra.html",
        itens=itens,
        separados=separados,
        tipo_estoque=tipo_estoque,
        total_sugerido=total_sugerido,
        total_critico=total_critico
    )


def montar_dados_ordem_compra(tipo_estoque=""):
    itens = buscar_itens_ordem_compra(tipo_estoque)

    cabecalhos = [
        "Produto",
        "Categoria",
        "Local",
        "Lote",
        "Atual",
        "Valido",
        "Vencido",
        "Perto do vencimento",
        "Minimo",
        "Padrao",
        "Sugerido comprar",
        "Acao sugerida",
        "Acao operacional",
        "Lotes vencidos",
        "Lotes perto do vencimento",
        "Motivo",
        "Status"
    ]

    linhas = []

    for item in itens:
        linhas.append([
            item["nome"],
            item["categoria"],
            item["tipo_estoque_nome"],
            item["lote"],
            item["quantidade_atual"],
            item["quantidade_valida"],
            item["quantidade_vencida"],
            item["quantidade_proxima_vencimento"],
            item["limite_alerta"],
            item["estoque_padrao"],
            item["sugestao"],
            item["acao_sugerida"],
            item["acao_operacional"]["mensagem"] if item["acao_operacional"] else "-",
            item["lotes_vencidos"],
            item["lotes_proximos_vencimento"],
            item["motivo_compra"],
            item["status"]
        ])

    titulo = "Ordem de Compra Sugerida"
    if tipo_estoque:
        titulo += f" - {nome_tipo_estoque(tipo_estoque)}"
    else:
        titulo += " - Todos os estoques"

    return titulo, cabecalhos, linhas


@app.route("/ordem_compra/exportar")
@login_obrigatorio
@licenca_obrigatoria
@estoque_obrigatorio
def exportar_ordem_compra():
    formato = request.args.get("formato", "xlsx")
    tipo_estoque = request.args.get("tipo_estoque", "").strip()

    if tipo_estoque not in ["", "almoxarifado", "farmacia_satelite", "carrinho_urgencia"]:
        tipo_estoque = ""

    if formato != "xlsx":
        flash("A ordem de compra agora é exportada apenas em Excel.", "erro")
        return redirect(url_for("ordem_compra"))

    titulo, cabecalhos, linhas = montar_dados_ordem_compra(tipo_estoque)

    data = datetime.now().strftime("%Y-%m-%d")
    sufixo = tipo_estoque.replace("_", "-") if tipo_estoque else "geral"
    nome_arquivo = f"ordem_compra_{sufixo}_{data}.{formato}"

    return gerar_excel(titulo, cabecalhos, linhas, nome_arquivo)

def obter_configuracao_alerta():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM configuracoes_alerta ORDER BY id LIMIT 1")
    config = cursor.fetchone()
    conn.close()
    return config


def registrar_historico_alerta(tipo_alerta, canal, destino, conteudo, status, erro=None):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO historico_alertas (
            tipo_alerta, canal, destino, conteudo, status, erro
        ) VALUES (%s, %s, %s, %s, %s, %s)
    """, (
        tipo_alerta,
        canal,
        destino,
        conteudo,
        status,
        erro
    ))
    conn.commit()
    conn.close()


def buscar_produtos_vencendo_para_alerta():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT produtos.*, categorias.nome AS categoria_nome
        FROM produtos
        JOIN categorias ON categorias.id = produtos.categoria_id
        ORDER BY produtos.tipo_estoque, produtos.nome
    """)
    produtos = cursor.fetchall()
    conn.close()

    itens = []

    for produto in produtos:
        status = calcular_status(produto)
        texto = status["texto"].lower()

        if "vence" in texto or "vencido" in texto:
            itens.append({
                "nome": produto["nome"],
                "categoria": produto["categoria_nome"],
                "local": nome_tipo_estoque(produto["tipo_estoque"]),
                "vencimento": formatar_data(produto["data_vencimento"]),
                "status": status["texto"]
            })

    return itens


def buscar_estoque_critico_para_alerta():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT produtos.*, categorias.nome AS categoria_nome
        FROM produtos
        JOIN categorias ON categorias.id = produtos.categoria_id
        WHERE produtos.quantidade_atual <= produtos.limite_alerta
        ORDER BY produtos.tipo_estoque, produtos.nome
    """)
    produtos = cursor.fetchall()
    conn.close()

    itens = []

    for produto in produtos:
        itens.append({
            "nome": produto["nome"],
            "categoria": produto["categoria_nome"],
            "local": nome_tipo_estoque(produto["tipo_estoque"]),
            "atual": produto["quantidade_atual"],
            "minimo": produto["limite_alerta"],
            "padrao": produto["estoque_padrao"]
        })

    return itens


def montar_mensagem_alertas():
    config = obter_configuracao_alerta()
    partes = []
    tipos_gerados = []

    if config and config["alertar_vencimentos"]:
        vencimentos = buscar_produtos_vencendo_para_alerta()

        if vencimentos:
            tipos_gerados.append("vencimentos")
            partes.append("Produtos proximos do vencimento ou vencidos:")

            for item in vencimentos[:20]:
                partes.append(
                    f"• {item['nome']} | {item['local']} | {item['status']} | venc.: {item['vencimento']}"
                )

    if config and config["alertar_estoque"]:
        estoque = buscar_estoque_critico_para_alerta()

        if estoque:
            tipos_gerados.append("estoque")
            partes.append("")
            partes.append("🚨 Estoque crítico/baixo:")

            for item in estoque[:20]:
                partes.append(
                    f"• {item['nome']} | {item['local']} | atual: {item['atual']} | mínimo: {item['minimo']}"
                )

    if config and config["alertar_ordem_compra"]:
        ordem = buscar_itens_ordem_compra()

        if ordem:
            tipos_gerados.append("ordem_compra")
            partes.append("")
            partes.append("ðŸ“¦ Ordem de compra sugerida:")

            for item in ordem[:20]:
                partes.append(
                    f"• {item['nome']} | {item['tipo_estoque_nome']} | comprar: {item['sugestao']}"
                )

    if not partes:
        return None, []

    mensagem = "Controle Oftalmo - Alertas automaticos\\n\\n" + "\\n".join(partes)
    return mensagem, tipos_gerados


def enviar_email_alerta(destino, assunto, mensagem):
    smtp_host = os.environ.get("SMTP_HOST")
    smtp_port = int(os.environ.get("SMTP_PORT", "587"))
    smtp_user = os.environ.get("SMTP_USER")
    smtp_password = os.environ.get("SMTP_PASSWORD")
    smtp_from = os.environ.get("SMTP_FROM", smtp_user)

    if not smtp_host or not smtp_user or not smtp_password:
        raise RuntimeError("SMTP não configurado. Configure SMTP_HOST, SMTP_USER e SMTP_PASSWORD.")

    email = EmailMessage()
    email["Subject"] = assunto
    email["From"] = smtp_from
    email["To"] = destino
    email.set_content(mensagem)

    with smtplib.SMTP(smtp_host, smtp_port) as server:
        server.starttls()
        server.login(smtp_user, smtp_password)
        server.send_message(email)


def enviar_whatsapp_alerta(telefone, mensagem):
    """
    Envio de WhatsApp por API.

    Modo 1 - UltraMsg:
    Configure no .env / Render:
    ULTRAMSG_INSTANCE_ID=instance000000
    ULTRAMSG_TOKEN=seu_token
    WHATSAPP_PHONE=5516999999999

    Modo 2 - Webhook generico:
    WHATSAPP_WEBHOOK_URL=https://sua-api.com/send
    WHATSAPP_TOKEN=token_opcional
    """

    ultra_instance = os.environ.get("ULTRAMSG_INSTANCE_ID")
    ultra_token = os.environ.get("ULTRAMSG_TOKEN")

    if ultra_instance and ultra_token:
        url = f"https://api.ultramsg.com/{ultra_instance}/messages/chat"
        payload = {
            "token": ultra_token,
            "to": telefone,
            "body": mensagem
        }

        resposta = requests.post(url, data=payload, timeout=25)
        resposta.raise_for_status()
        return True

    webhook_url = os.environ.get("WHATSAPP_WEBHOOK_URL")
    token = os.environ.get("WHATSAPP_TOKEN")

    if not webhook_url:
        raise RuntimeError("WhatsApp não configurado. Configure ULTRAMSG_INSTANCE_ID + ULTRAMSG_TOKEN ou WHATSAPP_WEBHOOK_URL.")

    payload = {
        "telefone": telefone,
        "mensagem": mensagem
    }

    headers = {}

    if token:
        headers["Authorization"] = f"Bearer {token}"

    resposta = requests.post(webhook_url, json=payload, headers=headers, timeout=25)
    resposta.raise_for_status()
    return True


def executar_envio_alertas(manual=True):
    config = obter_configuracao_alerta()

    if not config:
        return False, "Configuração de alerta não encontrada."

    mensagem, tipos = montar_mensagem_alertas()

    if not mensagem:
        return True, "Nenhum alerta encontrado no momento."

    assunto = "Controle Oftalmo - Alertas automaticos"
    enviados = []

    tipo_alerta = ",".join(tipos) if tipos else "geral"

    if config["usar_email"] and config["email_destino"]:
        try:
            enviar_email_alerta(config["email_destino"], assunto, mensagem)
            registrar_historico_alerta(tipo_alerta, "email", config["email_destino"], mensagem, "enviado")
            enviados.append("e-mail")
        except Exception as erro:
            registrar_historico_alerta(tipo_alerta, "email", config["email_destino"], mensagem, "erro", str(erro))
            if manual:
                flash(f"Erro ao enviar e-mail: {erro}", "erro")

    if config["usar_whatsapp"] and config["telefone_whatsapp"]:
        try:
            enviar_whatsapp_alerta(config["telefone_whatsapp"], mensagem)
            registrar_historico_alerta(tipo_alerta, "whatsapp", config["telefone_whatsapp"], mensagem, "enviado")
            enviados.append("WhatsApp")
        except Exception as erro:
            registrar_historico_alerta(tipo_alerta, "whatsapp", config["telefone_whatsapp"], mensagem, "erro", str(erro))
            if manual:
                flash(f"Erro ao enviar WhatsApp: {erro}", "erro")

    if enviados:
        return True, f"Alertas enviados por: {', '.join(enviados)}."

    return False, "Nenhum canal de envio ativo/configurado."


def garantir_tabela_auditoria_scanner():
    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS auditoria_scanner (
            id SERIAL PRIMARY KEY,
            codigo_barras TEXT,
            produto_id INTEGER,
            produto_nome TEXT,
            tipo_acao TEXT,
            quantidade INTEGER,
            usuario TEXT,
            ip_origem TEXT,
            navegador TEXT,
            status TEXT,
            detalhe TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    conn.commit()
    conn.close()


def registrar_auditoria_scanner(codigo_barras, produto_id, produto_nome, tipo_acao, quantidade, status, detalhe):
    garantir_tabela_auditoria_scanner()

    usuario = session.get("usuario_nome") or session.get("usuario_login") or "Admin"
    ip_origem = request.headers.get("X-Forwarded-For", request.remote_addr)
    navegador = request.headers.get("User-Agent", "")[:250]

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO auditoria_scanner (
            codigo_barras, produto_id, produto_nome, tipo_acao, quantidade,
            usuario, ip_origem, navegador, status, detalhe
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        codigo_barras,
        produto_id,
        produto_nome,
        tipo_acao,
        quantidade,
        usuario,
        ip_origem,
        navegador,
        status,
        detalhe
    ))

    conn.commit()
    conn.close()


def montar_dashboard_tempo_real():
    garantir_tabela_auditoria_scanner()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT COUNT(*) AS total
        FROM produtos
    """)
    total_produtos = cursor.fetchone()["total"]

    cursor.execute("""
        SELECT COUNT(*) AS total
        FROM (
            SELECT
                pe.id,
                pe.limite_alerta,
                COALESCE(SUM(l.quantidade_atual) FILTER (WHERE l.ativo = TRUE), 0) AS quantidade_atual
            FROM produto_estoques pe
            LEFT JOIN lotes l ON l.produto_estoque_id = pe.id
            WHERE pe.ativo = TRUE
            GROUP BY pe.id, pe.limite_alerta
        ) estoque
        WHERE quantidade_atual <= limite_alerta
    """)
    estoque_baixo = cursor.fetchone()["total"]

    cursor.execute("""
        SELECT COUNT(*) AS total
        FROM (
            SELECT
                pe.id,
                COALESCE(SUM(l.quantidade_atual) FILTER (WHERE l.ativo = TRUE), 0) AS quantidade_atual
            FROM produto_estoques pe
            LEFT JOIN lotes l ON l.produto_estoque_id = pe.id
            WHERE pe.ativo = TRUE
            GROUP BY pe.id
        ) estoque
        WHERE quantidade_atual <= 0
    """)
    estoque_zerado = cursor.fetchone()["total"]

    cursor.execute("""
        SELECT COUNT(*) AS total
        FROM produto_estoques
        WHERE tipo_estoque = 'almoxarifado'
          AND ativo = TRUE
    """)
    total_almoxarifado = cursor.fetchone()["total"]

    cursor.execute("""
        SELECT COUNT(*) AS total
        FROM produto_estoques
        WHERE tipo_estoque = 'farmacia_satelite'
          AND ativo = TRUE
    """)
    total_farmacia = cursor.fetchone()["total"]

    cursor.execute("""
        SELECT COALESCE(SUM(l.quantidade_atual), 0) AS total
        FROM produto_estoques pe
        JOIN lotes l ON l.produto_estoque_id = pe.id
        WHERE pe.tipo_estoque = 'almoxarifado'
          AND pe.ativo = TRUE
          AND l.ativo = TRUE
    """)
    unidades_almoxarifado = cursor.fetchone()["total"]

    cursor.execute("""
        SELECT COALESCE(SUM(l.quantidade_atual), 0) AS total
        FROM produto_estoques pe
        JOIN lotes l ON l.produto_estoque_id = pe.id
        WHERE pe.tipo_estoque = 'farmacia_satelite'
          AND pe.ativo = TRUE
          AND l.ativo = TRUE
    """)
    unidades_farmacia = cursor.fetchone()["total"]

    cursor.execute("""
        SELECT *
        FROM movimentacoes
        ORDER BY data_movimentacao DESC
        LIMIT 10
    """)
    ultimas_movimentacoes = cursor.fetchall()

    cursor.execute("""
        SELECT *
        FROM auditoria_scanner
        ORDER BY criado_em DESC
        LIMIT 10
    """)
    ultimas_bipagens = cursor.fetchall()

    cursor.execute("""
        SELECT produto_nome, SUM(quantidade) AS total
        FROM movimentacoes
        WHERE tipo_movimentacao IN ('retirada', 'saida')
        GROUP BY produto_nome
        ORDER BY total DESC
        LIMIT 5
    """)
    produtos_mais_movimentados = cursor.fetchall()

    cursor.execute("""
        SELECT
            p.*,
            c.nome AS categoria_nome,
            pe.tipo_estoque,
            pe.limite_alerta,
            COALESCE(SUM(l.quantidade_atual) FILTER (WHERE l.ativo = TRUE), 0) AS quantidade_atual
        FROM produto_estoques pe
        JOIN produtos p ON p.id = pe.produto_id
        LEFT JOIN categorias c ON c.id = p.categoria_id
        LEFT JOIN lotes l ON l.produto_estoque_id = pe.id
        WHERE pe.ativo = TRUE
        GROUP BY p.id, c.nome, pe.id, pe.tipo_estoque, pe.limite_alerta
        HAVING COALESCE(SUM(l.quantidade_atual) FILTER (WHERE l.ativo = TRUE), 0) <= pe.limite_alerta
        ORDER BY quantidade_atual ASC
        LIMIT 8
    """)
    produtos_criticos_tempo_real = cursor.fetchall()

    cursor.execute("""
        SELECT COUNT(*) AS total
        FROM auditoria_scanner
        WHERE status = 'sucesso'
          AND criado_em::date = CURRENT_DATE
    """)
    scanner_sucesso_hoje = cursor.fetchone()["total"]

    cursor.execute("""
        SELECT COUNT(*) AS total
        FROM auditoria_scanner
        WHERE status = 'erro'
          AND criado_em::date = CURRENT_DATE
    """)
    scanner_erros_hoje = cursor.fetchone()["total"]

    conn.close()

    return {
        "total_produtos": total_produtos,
        "estoque_baixo": estoque_baixo,
        "estoque_zerado": estoque_zerado,
        "total_almoxarifado": total_almoxarifado,
        "total_farmacia": total_farmacia,
        "unidades_almoxarifado": unidades_almoxarifado,
        "unidades_farmacia": unidades_farmacia,
        "ultimas_movimentacoes": ultimas_movimentacoes,
        "ultimas_bipagens": ultimas_bipagens,
        "produtos_mais_movimentados": produtos_mais_movimentados,
        "produtos_criticos": produtos_criticos_tempo_real,
        "scanner_sucesso_hoje": scanner_sucesso_hoje,
        "scanner_erros_hoje": scanner_erros_hoje
    }



def pode_enviar_whatsapp_automatico(config):
    if not config:
        return False

    if not config.get("ultimo_envio_whatsapp"):
        return True

    intervalo = config.get("intervalo_minutos") or 720
    ultimo = config["ultimo_envio_whatsapp"]

    if isinstance(ultimo, str):
        try:
            ultimo = datetime.fromisoformat(ultimo)
        except Exception:
            return True

    return datetime.now() - ultimo >= timedelta(minutes=int(intervalo))


def marcar_envio_whatsapp_realizado():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE configuracoes_alerta
        SET ultimo_envio_whatsapp = NOW()
        WHERE id = (SELECT id FROM configuracoes_alerta ORDER BY id LIMIT 1)
    """)
    conn.commit()
    conn.close()


def executar_alerta_whatsapp_automatico(forcar=False):
    config = obter_configuracao_alerta()

    if not config:
        return False, "Configuração de alerta não encontrada."

    if not config["usar_whatsapp"]:
        return False, "WhatsApp automático está desativado."

    telefone = config["telefone_whatsapp"]

    if not telefone:
        return False, "Telefone de WhatsApp não configurado."

    if not forcar and not pode_enviar_whatsapp_automatico(config):
        return True, "Envio ignorado para evitar mensagens repetidas no intervalo configurado."

    mensagem, tipos = montar_mensagem_alertas()

    if not mensagem:
        return True, "Nenhum alerta crítico encontrado para envio."

    try:
        enviar_whatsapp_alerta(telefone, mensagem)
        registrar_historico_alerta(
            ",".join(tipos) if tipos else "geral",
            "whatsapp",
            telefone,
            mensagem,
            "enviado",
            None
        )
        marcar_envio_whatsapp_realizado()
        return True, "Alerta enviado por WhatsApp com sucesso."
    except Exception as erro:
        registrar_historico_alerta(
            ",".join(tipos) if tipos else "geral",
            "whatsapp",
            telefone,
            mensagem,
            "erro",
            str(erro)
        )
        return False, f"Erro ao enviar WhatsApp: {erro}"

@app.route("/dashboard_tempo_real")
@login_obrigatorio
@licenca_obrigatoria
@estoque_obrigatorio
@admin_obrigatorio
def dashboard_tempo_real():
    dados = montar_dashboard_tempo_real()
    return render_template("dashboard_tempo_real.html", dados=dados)


@app.route("/usuarios")
@login_obrigatorio
@admin_obrigatorio
def usuarios():
    garantir_tabela_usuarios()
    usuarios_lista = listar_usuarios()
    return render_template("usuarios.html", usuarios=usuarios_lista)


@app.route("/usuarios/novo", methods=["GET", "POST"])
@login_obrigatorio
@admin_obrigatorio
def novo_usuario():
    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        usuario = request.form.get("usuario", "").strip().lower()
        senha = request.form.get("senha", "")
        perfil = request.form.get("perfil", "estoque")

        if not nome or not usuario or not senha:
            flash("Preencha nome, usuário e senha.", "erro")
            return render_template("usuario_form.html", usuario=None, titulo="Novo Usuário")

        conn = conectar()
        cursor = conn.cursor()

        try:
            cursor.execute("""
                INSERT INTO usuarios (nome, usuario, senha_hash, perfil, ativo)
                VALUES (%s, %s, %s, %s, TRUE)
            """, (
                nome,
                usuario,
                generate_password_hash(senha),
                perfil
            ))
            conn.commit()
            flash("Usuário cadastrado com sucesso.", "sucesso")
        except Exception:
            conn.rollback()
            flash("Não foi possível cadastrar. Verifique se o usuário já existe.", "erro")
        finally:
            conn.close()

        return redirect(url_for("usuarios"))

    return render_template("usuario_form.html", usuario=None, titulo="Novo Usuário")


@app.route("/usuarios/editar/<int:id>", methods=["GET", "POST"])
@login_obrigatorio
@admin_obrigatorio
def editar_usuario(id):
    garantir_tabela_usuarios()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM usuarios WHERE id = %s", (id,))
    usuario = cursor.fetchone()

    if not usuario:
        conn.close()
        flash("Usuário não encontrado.", "erro")
        return redirect(url_for("usuarios"))

    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        perfil = request.form.get("perfil", "estoque")
        ativo = True if request.form.get("ativo") == "on" else False
        nova_senha = request.form.get("senha", "")

        if nova_senha:
            cursor.execute("""
                UPDATE usuarios
                SET nome = %s,
                    perfil = %s,
                    ativo = %s,
                    senha_hash = %s
                WHERE id = %s
            """, (
                nome,
                perfil,
                ativo,
                generate_password_hash(nova_senha),
                id
            ))
        else:
            cursor.execute("""
                UPDATE usuarios
                SET nome = %s,
                    perfil = %s,
                    ativo = %s
                WHERE id = %s
            """, (
                nome,
                perfil,
                ativo,
                id
            ))

        conn.commit()
        conn.close()

        flash("Usuário atualizado com sucesso.", "sucesso")
        return redirect(url_for("usuarios"))

    conn.close()
    return render_template("usuario_form.html", usuario=usuario, titulo="Editar Usuário")


@app.route("/usuarios/desativar/<int:id>", methods=["POST"])
@login_obrigatorio
@admin_obrigatorio
def desativar_usuario(id):
    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE usuarios
        SET ativo = FALSE
        WHERE id = %s
    """, (id,))

    conn.commit()
    conn.close()

    flash("Usuário desativado.", "sucesso")
    return redirect(url_for("usuarios"))


@app.route("/codigo_barras", methods=["GET", "POST"])
@login_obrigatorio
@licenca_obrigatoria
@estoque_obrigatorio
def codigo_barras():
    codigo = request.values.get("codigo", "").strip()
    tipo_estoque = request.values.get("tipo_estoque", "almoxarifado").strip()
    if tipo_estoque not in ESTOQUES:
        tipo_estoque = "almoxarifado"

    produto = None
    status = None

    if codigo:
        conn = conectar()
        cursor = conn.cursor()
        produto = buscar_lote_por_codigo_scanner(cursor, codigo, tipo_estoque)
        conn.close()

        if produto:
            status = calcular_status(produto)
        else:
            flash("Nenhum lote disponível e válido encontrado para este código no estoque escolhido.", "erro")

    garantir_tabela_auditoria_scanner()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT *
        FROM movimentacoes
        WHERE observacao ILIKE %s
        ORDER BY data_movimentacao DESC
        LIMIT 8
    """, ("%scanner%",))
    ultimos_scanner = cursor.fetchall()

    cursor.execute("""
        SELECT *
        FROM auditoria_scanner
        ORDER BY criado_em DESC
        LIMIT 10
    """)
    auditoria_scanner = cursor.fetchall()

    total_bipagens = len(ultimos_scanner)
    total_unidades = sum(item["quantidade"] for item in ultimos_scanner) if ultimos_scanner else 0

    estoque_baixo_scanner = len([
        produto for produto in listar_produtos_resumo_estoque()
        if produto["quantidade_atual"] <= produto["limite_alerta"]
    ])

    conn.close()

    return render_template(
        "codigo_barras.html",
        codigo=codigo,
        tipo_estoque=tipo_estoque,
        produto=produto,
        status=status,
        ultimos_scanner=ultimos_scanner,
        auditoria_scanner=auditoria_scanner,
        total_bipagens=total_bipagens,
        total_unidades=total_unidades,
        estoque_baixo_scanner=estoque_baixo_scanner
    )


@app.route("/scanner_baixa_rapida", methods=["POST"])
@login_obrigatorio
@alteracao_permitida
@estoque_obrigatorio
def scanner_baixa_rapida():
    codigo = request.form.get("codigo", "").strip()
    tipo_estoque = request.form.get("tipo_estoque", "almoxarifado").strip()
    observacao = request.form.get("observacao", "Baixa por Scanner Operacional").strip() or "Baixa por Scanner Operacional"

    if tipo_estoque not in ESTOQUES:
        tipo_estoque = "almoxarifado"

    try:
        quantidade = int(request.form.get("quantidade") or 1)
    except ValueError:
        quantidade = 1

    if not codigo:
        flash("Bipe ou informe um código de barras.", "erro")
        return redirect(url_for("codigo_barras", modo="continuo", tipo_estoque=tipo_estoque))

    conn = conectar()
    cursor = conn.cursor()
    produto = buscar_lote_por_codigo_scanner(cursor, codigo, tipo_estoque)

    if not produto:
        conn.close()
        registrar_auditoria_scanner(
            codigo,
            None,
            None,
            "baixa_rapida",
            quantidade,
            "erro",
            f"Nenhum lote disponível em {nome_tipo_estoque(tipo_estoque)}"
        )
        flash("Nenhum lote disponível e válido para baixa rápida neste estoque.", "erro")
        return redirect(url_for("codigo_barras", codigo=codigo, modo="continuo", tipo_estoque=tipo_estoque))

    if produto_esta_vencido(produto):
        conn.close()
        registrar_auditoria_scanner(
            codigo,
            produto["id"],
            produto["nome"],
            "baixa_rapida",
            quantidade,
            "erro",
            "Produto vencido bloqueado"
        )
        flash("Produto vencido bloqueado para baixa por scanner.", "erro")
        return redirect(url_for("codigo_barras", codigo=codigo, modo="continuo", tipo_estoque=tipo_estoque))


    if produto["quantidade_atual"] < quantidade:
        conn.close()
        registrar_auditoria_scanner(
            codigo,
            produto["id"],
            produto["nome"],
            "baixa_rapida",
            quantidade,
            "erro",
            "Estoque insuficiente"
        )
        flash("Estoque insuficiente para baixa rápida.", "erro")
        return redirect(url_for("codigo_barras", codigo=codigo, modo="continuo", tipo_estoque=tipo_estoque))

    estoque_anterior = produto["quantidade_atual"]
    estoque_atual = estoque_anterior - quantidade

    cursor.execute("""
        UPDATE lotes
        SET quantidade_atual = %s
        WHERE id = %s
    """, (estoque_atual, produto["lote_id"]))

    registrar_movimentacao_lote(
        cursor,
        produto,
        "saida",
        quantidade,
        estoque_anterior,
        estoque_atual,
        observacao,
        produto["tipo_estoque"],
        None
    )

    conn.commit()
    conn.close()

    registrar_auditoria_scanner(
        codigo,
        produto["produto_id"],
        produto["nome"],
        "baixa_rapida",
        quantidade,
        "sucesso",
        f"Lote {produto['numero_lote'] or '-'} baixado por scanner em {nome_tipo_estoque(tipo_estoque)}"
    )

    flash(f"Baixa rápida realizada: {produto['nome']} (-{quantidade}) no lote {produto['numero_lote'] or '-'}.", "sucesso")
    return redirect(url_for("codigo_barras", modo="continuo", scanner_status="sucesso", tipo_estoque=tipo_estoque))


@app.route("/alertas_visual")
@login_obrigatorio
@licenca_obrigatoria
@estoque_obrigatorio
def alertas_visual():
    mensagem = ""
    tipos = []

    try:
        mensagem, tipos = montar_mensagem_alertas()
    except Exception:
        mensagem = "Não foi possível montar os alertas no momento."
        tipos = []

    return render_template(
        "alertas_visual.html",
        mensagem_alertas=mensagem,
        tipos_alertas=tipos
    )

@app.route("/config_alertas", methods=["GET", "POST"])
@login_obrigatorio
@alteracao_permitida
@admin_obrigatorio
def config_alertas():
    config = obter_configuracao_alerta()

    if request.method == "POST":
        email_destino = request.form.get("email_destino", "").strip()
        telefone_whatsapp = request.form.get("telefone_whatsapp", "").strip()
        intervalo_minutos = int(request.form.get("intervalo_minutos") or 720)
        usar_email = bool(request.form.get("usar_email"))
        usar_whatsapp = bool(request.form.get("usar_whatsapp"))
        alertar_vencimentos = bool(request.form.get("alertar_vencimentos"))
        alertar_estoque = bool(request.form.get("alertar_estoque"))
        alertar_ordem_compra = bool(request.form.get("alertar_ordem_compra"))
        hora_envio = request.form.get("hora_envio", "08:00").strip() or "08:00"

        conn = conectar()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE configuracoes_alerta SET
                email_destino = %s,
                telefone_whatsapp = %s,
                    intervalo_minutos = %s,
                usar_email = %s,
                usar_whatsapp = %s,
                alertar_vencimentos = %s,
                alertar_estoque = %s,
                alertar_ordem_compra = %s,
                hora_envio = %s
            WHERE id = %s
        """, (
            email_destino,
            telefone_whatsapp,
                    intervalo_minutos,
            usar_email,
            usar_whatsapp,
            alertar_vencimentos,
            alertar_estoque,
            alertar_ordem_compra,
            hora_envio,
            config["id"]
        ))
        conn.commit()
        conn.close()

        flash("Configurações de alerta salvas com sucesso.", "sucesso")
        return redirect(url_for("config_alertas"))

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM historico_alertas ORDER BY criado_em DESC LIMIT 20")
    historico = cursor.fetchall()
    conn.close()

    #mensagem_preview, _ = montar_mensagem_alertas()
    mensagem_preview = ""

    return render_template(
        "config_alertas.html",
        config=config,
        historico=historico,
       mensagem_preview=mensagem_preview
    )


@app.route("/config_alertas/testar", methods=["POST"])
@login_obrigatorio
@alteracao_permitida
@admin_obrigatorio
def testar_alertas():
    sucesso, mensagem = executar_envio_alertas(manual=True)

    if sucesso:
        flash(mensagem, "sucesso")
    else:
        flash(mensagem, "erro")

    return redirect(url_for("config_alertas"))


def senha_padrao_ativa():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT senha_hash FROM admin WHERE usuario = %s", ("admin",))
    admin = cursor.fetchone()
    conn.close()

    if not admin:
        return False

    return check_password_hash(admin["senha_hash"], "admin123")


@app.route("/configuracoes")
@login_obrigatorio
@licenca_obrigatoria
@admin_obrigatorio
def configuracoes():
    return render_template("configuracoes.html", senha_padrao=senha_padrao_ativa())


@app.route("/configuracoes/alterar_senha", methods=["POST"])
@login_obrigatorio
@alteracao_permitida
@admin_obrigatorio
def alterar_senha():
    senha_atual = request.form.get("senha_atual", "").strip()
    nova_senha = request.form.get("nova_senha", "").strip()
    confirmar_senha = request.form.get("confirmar_senha", "").strip()

    if len(nova_senha) < 6:
        flash("A nova senha deve ter pelo menos 6 caracteres.", "erro")
        return redirect(url_for("configuracoes"))

    if nova_senha != confirmar_senha:
        flash("A confirmação da nova senha não confere.", "erro")
        return redirect(url_for("configuracoes"))

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM admin WHERE usuario = %s", (session.get("usuario", "admin"),))
    admin = cursor.fetchone()

    if not admin or not check_password_hash(admin["senha_hash"], senha_atual):
        conn.close()
        flash("Senha atual incorreta.", "erro")
        return redirect(url_for("configuracoes"))

    cursor.execute("UPDATE admin SET senha_hash = %s WHERE id = %s", (generate_password_hash(nova_senha), admin["id"]))
    conn.commit()
    conn.close()

    flash("Senha alterada com sucesso.", "sucesso")
    return redirect(url_for("configuracoes"))

@app.route("/painel_master_reset_7319", methods=["GET", "POST"])
@login_obrigatorio
@admin_obrigatorio
def reset_sistema():
    if request.method == "POST":
        conn = conectar()
        cursor = conn.cursor()

        cursor.execute("DELETE FROM movimentacoes")
        cursor.execute("DELETE FROM historico_alertas")
        cursor.execute("DELETE FROM produtos")
        cursor.execute("DELETE FROM categorias")
        cursor.execute("DELETE FROM auditoria_scanner")

        cursor.execute("ALTER SEQUENCE produtos_id_seq RESTART WITH 1")
        cursor.execute("ALTER SEQUENCE movimentacoes_id_seq RESTART WITH 1")
        cursor.execute("ALTER SEQUENCE categorias_id_seq RESTART WITH 1")
        cursor.execute("ALTER SEQUENCE historico_alertas_id_seq RESTART WITH 1")
        cursor.execute("ALTER SEQUENCE auditoria_scanner_id_seq RESTART WITH 1")

        conn.commit()
        conn.close()

        flash("Sistema zerado com sucesso.", "sucesso")
        return redirect(url_for("dashboard"))

    return """
    <h1>Zerar sistema</h1>
    <p>Isso vai apagar produtos, categorias, movimentações e histórico de alertas.</p>
    <form method="POST">
        <button type="submit" onclick="return confirm('Tem certeza que deseja zerar o sistema?')">
            Zerar sistema
        </button>
    </form>
    """

# Inicialização do banco para ambiente local e deploy
criar_banco()

iniciar_backup_automatico()

@app.route("/importar_estoque", methods=["GET", "POST"])
@login_obrigatorio
@alteracao_permitida
@estoque_obrigatorio
def importar_estoque():
    if request.method == "POST":
        arquivo = request.files.get("arquivo")

        if not arquivo:
            flash("Selecione uma planilha Excel.", "erro")
            return redirect(url_for("importar_estoque"))

        try:
            wb = load_workbook(arquivo)
            ws = wb.active

            conn = conectar()
            cursor = conn.cursor()

            importados = 0
            atualizados = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                nome, quantidade, lote, data_vencimento, categoria_nome, tipo_estoque, codigo_barras, estoque_padrao, limite_alerta, observacoes = row

                if not nome:
                    continue

                nome = str(nome).strip()
                lote = str(lote or "").strip()
                categoria_nome = str(categoria_nome or "Materiais").strip()
                tipo_estoque = str(tipo_estoque or "almoxarifado").strip()
                codigo_barras = str(codigo_barras or "").strip()
                observacoes = str(observacoes or "").strip()

                quantidade = int(quantidade or 0)
                estoque_padrao = int(estoque_padrao or quantidade)
                limite_alerta = int(limite_alerta or 1)

                cursor.execute("""
                    INSERT INTO categorias (nome)
                    VALUES (%s)
                    ON CONFLICT (nome) DO NOTHING
                """, (categoria_nome,))

                cursor.execute("SELECT id FROM categorias WHERE nome = %s", (categoria_nome,))
                categoria = cursor.fetchone()
                categoria_id = categoria["id"]

                cursor.execute("""
                    SELECT id, quantidade_atual
                    FROM produtos
                    WHERE LOWER(TRIM(nome)) = LOWER(TRIM(%s))
                      AND COALESCE(lote, '') = COALESCE(%s, '')
                      AND data_vencimento = %s
                      AND tipo_estoque = %s
                    LIMIT 1
                """, (nome, lote, data_vencimento, tipo_estoque))

                existente = cursor.fetchone()

                if existente:
                    nova_quantidade = existente["quantidade_atual"] + quantidade

                    cursor.execute("""
                        UPDATE produtos
                        SET quantidade_atual = %s
                        WHERE id = %s
                    """, (nova_quantidade, existente["id"]))

                    atualizados += 1

                else:
                    cursor.execute("""
                        INSERT INTO produtos (
                            nome, categoria_id, lote, codigo_barras,
                            data_vencimento, quantidade_atual, estoque_padrao,
                            limite_alerta, observacoes, tipo_estoque
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        nome, categoria_id, lote, codigo_barras,
                        data_vencimento, quantidade, estoque_padrao,
                        limite_alerta, observacoes, tipo_estoque
                    ))

                    importados += 1

            conn.commit()
            conn.close()

            flash(f"Importação concluída. Novos: {importados}. Atualizados: {atualizados}.", "sucesso")
            return redirect(url_for("produtos"))

        except Exception as erro:
            flash(f"Erro ao importar planilha: {erro}", "erro")
            return redirect(url_for("importar_estoque"))

    return render_template("importar_estoque.html")

@app.route("/backup/restaurar_local/<nome_arquivo>", methods=["POST"])
@login_obrigatorio
@admin_obrigatorio
def restaurar_backup_local(nome_arquivo):
    try:
        # Backup de segurança antes de restaurar
        gerar_backup_sistema()

        caminho = os.path.join(BACKUP_DIR, nome_arquivo)

        if not os.path.exists(caminho):
            flash("Arquivo de backup não encontrado.", "erro")
            return redirect(url_for("painel_backup"))

        with open(caminho, "r", encoding="utf-8") as arquivo:
            backup = json.load(arquivo)

        dados = backup.get("dados", {})

        conn = conectar()
        cursor = conn.cursor()

        cursor.execute("DELETE FROM movimentacoes")
        cursor.execute("DELETE FROM produtos")
        cursor.execute("DELETE FROM categorias")

        for categoria in dados.get("categorias", []):
            cursor.execute("""
                INSERT INTO categorias (id, nome, descricao, criado_em)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (id) DO NOTHING
            """, (
                categoria.get("id"),
                categoria.get("nome"),
                categoria.get("descricao"),
                categoria.get("criado_em")
            ))

        for produto in dados.get("produtos", []):
            cursor.execute("""
                INSERT INTO produtos (
                    id, nome, categoria_id, lote, codigo_barras,
                    data_vencimento, data_abertura, validade_apos_aberto_dias,
                    quantidade_atual, estoque_padrao, limite_alerta,
                    observacoes, tipo_estoque, criado_em
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (id) DO NOTHING
            """, (
                produto.get("id"),
                produto.get("nome"),
                produto.get("categoria_id"),
                produto.get("lote"),
                produto.get("codigo_barras"),
                produto.get("data_vencimento"),
                produto.get("data_abertura"),
                produto.get("validade_apos_aberto_dias"),
                produto.get("quantidade_atual"),
                produto.get("estoque_padrao"),
                produto.get("limite_alerta"),
                produto.get("observacoes"),
                produto.get("tipo_estoque", "almoxarifado"),
                produto.get("criado_em")
            ))

        for mov in dados.get("movimentacoes", []):
            cursor.execute("""
                INSERT INTO movimentacoes (
                    id, produto_id, produto_nome, tipo_movimentacao,
                    quantidade, estoque_anterior, estoque_atual,
                    observacao, tipo_estoque, estoque_origem,
                    estoque_destino, usuario_id, usuario_nome,
                    data_movimentacao
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (id) DO NOTHING
            """, (
                mov.get("id"),
                mov.get("produto_id"),
                mov.get("produto_nome"),
                mov.get("tipo_movimentacao"),
                mov.get("quantidade"),
                mov.get("estoque_anterior"),
                mov.get("estoque_atual"),
                mov.get("observacao"),
                mov.get("tipo_estoque"),
                mov.get("estoque_origem"),
                mov.get("estoque_destino"),
                mov.get("usuario_id"),
                mov.get("usuario_nome"),
                mov.get("data_movimentacao")
            ))

        cursor.execute("SELECT setval('categorias_id_seq', COALESCE((SELECT MAX(id) FROM categorias), 1))")
        cursor.execute("SELECT setval('produtos_id_seq', COALESCE((SELECT MAX(id) FROM produtos), 1))")
        cursor.execute("SELECT setval('movimentacoes_id_seq', COALESCE((SELECT MAX(id) FROM movimentacoes), 1))")

        conn.commit()
        conn.close()

        flash("Backup restaurado com sucesso. Um backup de segurança foi criado antes da restauração.", "sucesso")
        return redirect(url_for("produtos"))

    except Exception as erro:
        flash(f"Erro ao restaurar backup: {erro}", "erro")
        return redirect(url_for("painel_backup"))

@app.route("/backup_automatico_9182")
def backup_automatico_externo():
    try:
        caminho, upload_ok, upload_mensagem = gerar_backup_sistema()

        return {
            "status": "ok" if upload_ok else "erro",
            "mensagem": upload_mensagem,
            "arquivo": os.path.basename(caminho)
        }, 200 if upload_ok else 500

    except Exception as erro:
        return {
            "status": "erro",
            "mensagem": str(erro)
        }, 500

if __name__ == "__main__":
    app.run(debug=True)






